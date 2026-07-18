"""
SQLite database for paper trading trade persistence.

This module provides a simple, efficient SQLite-based database for storing
and retrieving paper trading trades. It supports:
- Saving individual trades
- Loading trades by various filters
- Statistics calculation
- Easy export to other formats
"""

from __future__ import annotations

import csv
import json
import logging
import sqlite3
import shutil
import time
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, List, Dict, Any, Callable, Set
from functools import lru_cache
import hashlib
from threading import Lock, Thread, Event
from contextlib import contextmanager
from concurrent.futures import ThreadPoolExecutor, as_completed
from queue import Queue, Empty

from .security import (
    DatabaseEncryption,
    AuthenticationManager,
    AuthorizationManager,
    DataMasker,
    AuthMethod,
    Role,
    User,
    MaskingRule,
)

log = logging.getLogger(__name__)


@dataclass
class TradeRecord:
    """Record of a single trade from the database."""
    id: int
    market_slug: str
    market_id: str
    side: str
    entry_price: float
    exit_price: Optional[float]
    amount: float
    shares: float
    fee: float
    outcome: Optional[str]  # "WON" | "LOST" | "CLOSED" | None
    pnl: float
    timestamp: datetime
    market_session: Optional[str] = None  # "london" | "new_york" | "asia" | "sydney" | None
    
    def to_dict(self) -> dict:
        """Convert to dictionary for serialization."""
        return {
            "id": self.id,
            "market_slug": self.market_slug,
            "market_id": self.market_id,
            "side": self.side,
            "entry_price": self.entry_price,
            "exit_price": self.exit_price,
            "amount": self.amount,
            "shares": self.shares,
            "fee": self.fee,
            "outcome": self.outcome,
            "pnl": self.pnl,
            "timestamp": self.timestamp.isoformat(),
            "market_session": self.market_session,
        }


@dataclass
class TradeStatistics:
    """Statistics summary for all trades."""
    total_trades: int
    wins: int
    losses: int
    win_rate: float
    total_pnl: float
    total_fees: float
    avg_entry_price: float
    avg_pnl_per_trade: float


@dataclass
class DatabaseMetrics:
    """Database performance and health metrics."""
    total_trades: int
    database_size_bytes: int
    cache_hit_rate: float
    cache_size: int
    query_count: int
    slow_query_count: int
    avg_query_time_ms: float
    connection_pool_size: int
    wal_enabled: bool
    last_optimization: Optional[datetime]
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for serialization."""
        return {
            "total_trades": self.total_trades,
            "database_size_bytes": self.database_size_bytes,
            "database_size_mb": round(self.database_size_bytes / (1024 * 1024), 2),
            "cache_hit_rate": round(self.cache_hit_rate * 100, 2),
            "cache_size": self.cache_size,
            "query_count": self.query_count,
            "slow_query_count": self.slow_query_count,
            "avg_query_time_ms": round(self.avg_query_time_ms, 2),
            "connection_pool_size": self.connection_pool_size,
            "wal_enabled": self.wal_enabled,
            "last_optimization": self.last_optimization.isoformat() if self.last_optimization else None,
        }


@dataclass
class LogEntry:
    """Structured log entry with correlation ID."""
    correlation_id: str
    timestamp: datetime
    level: str
    message: str
    operation: Optional[str]
    duration_ms: Optional[float]
    metadata: Dict[str, Any]
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for serialization."""
        return {
            "correlation_id": self.correlation_id,
            "timestamp": self.timestamp.isoformat(),
            "level": self.level,
            "message": self.message,
            "operation": self.operation,
            "duration_ms": self.duration_ms,
            "metadata": self.metadata,
        }


@dataclass
class AlertRule:
    """Alert rule definition."""
    name: str
    metric: str
    threshold: float
    comparison: str  # "gt", "lt", "eq", "gte", "lte"
    enabled: bool
    callback: Optional[Callable[[str, float, float], None]]
    last_triggered: Optional[datetime]
    trigger_count: int


class TradeDatabase:
    """
    SQLite database for paper trading trade persistence.
    
    This class provides a simple interface for saving and loading
    paper trading trades to a SQLite database file.
    
    Parameters
    ----------
    db_path : str or Path
        Path to the SQLite database file. Will be created if it doesn't exist.
    
    Example
    -------
    >>> db = TradeDatabase("trades.db")
    >>> db.save_trade(
    ...     market_slug="btc-updown-5m-1751234700",
    ...     market_id="abc123",
    ...     side="UP",
    ...     entry_price=0.92,
    ...     exit_price=None,
    ...     amount=10.0,
    ...     shares=10.5,
    ...     fee=0.2,
    ...     outcome="WON",
    ...     pnl=5.3,
    ...     timestamp=datetime.now(timezone.utc)
    ... )
    >>> trades = db.load_all_trades()
    """
    
    def __init__(self, db_path: str | Path, enable_wal: bool = True, enable_cache: bool = True):
        """
        Initialize the database.
        
        Parameters
        ----------
        db_path : str or Path
            Path to the SQLite database file.
        enable_wal : bool, optional
            Enable WAL (Write-Ahead Logging) mode for better concurrency (default: True).
        enable_cache : bool, optional
            Enable query result caching (default: True).
        """
        self.db_path = Path(db_path)
        self._conn: Optional[sqlite3.Connection] = None
        self._enable_wal = enable_wal
        self._cache_enabled = enable_cache
        self._query_cache: Dict[str, List[TradeRecord]] = {}
        self._cache_max_size = 100
        self._cache_ttl: Dict[str, float] = {}  # Cache entry TTL
        self._default_cache_ttl = 300.0  # 5 minutes default TTL
        
        # Prepared statement cache
        self._prepared_statements: Dict[str, sqlite3.Statement] = {}
        self._statement_cache_max_size = 50
        self._statement_lock = Lock()
        
        # Event hooks for real-time synchronization
        self._trade_saved_hooks: List[Callable[[TradeRecord], None]] = []
        self._trade_updated_hooks: List[Callable[[int, Dict[str, Any]], None]] = []
        self._trade_deleted_hooks: List[Callable[[int], None]] = []
        self._hooks_lock = Lock()
        
        # Streaming state
        self._streaming_enabled = False
        
        # Monitoring and observability
        self._query_count = 0
        self._slow_query_count = 0
        self._query_times: List[float] = []
        self._max_query_times = 1000  # Keep last 1000 query times
        self._slow_query_threshold_ms = 1000.0  # 1 second
        self._log_entries: List[LogEntry] = []
        self._max_log_entries = 10000  # Keep last 10000 log entries
        self._log_lock = Lock()
        self._alert_rules: Dict[str, AlertRule] = {}
        self._alert_lock = Lock()
        self._current_correlation_id: Optional[str] = None
        self._correlation_lock = Lock()
        self._last_optimization: Optional[datetime] = None
        self._cache_hits = 0
        self._cache_misses = 0
        
        # Security features
        self._encryption: Optional[DatabaseEncryption] = None
        self._auth_manager = AuthenticationManager()
        self._authz_manager = AuthorizationManager()
        self._data_masker = DataMasker()
        self._current_user_id: Optional[str] = None
        self._current_roles: Set[str] = set()
        self._encryption_fields: List[str] = []  # Fields to encrypt
        
        # Connection pool for thread safety
        self._connection_pool: Queue[sqlite3.Connection] = Queue(maxsize=5)
        self._pool_lock = Lock()
        self._max_pool_size = 5
        self._pool_created = 0
        
        # Background optimization scheduler
        self._optimization_thread: Optional[Thread] = None
        self._optimization_stop_event = Event()
        self._optimization_interval = 3600  # 1 hour
        
        self._initialize_db()
    
    def _get_connection(self) -> sqlite3.Connection:
        """Get or create database connection from pool."""
        try:
            # Try to get connection from pool
            conn = self._connection_pool.get_nowait()
            return conn
        except Empty:
            # Pool is empty, create new connection
            with self._pool_lock:
                if self._pool_created < self._max_pool_size:
                    conn = self._create_connection()
                    self._pool_created += 1
                    return conn
                else:
                    # Pool is full, wait for available connection
                    return self._connection_pool.get(timeout=5)
    
    def _create_connection(self) -> sqlite3.Connection:
        """Create a new database connection with optimizations."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        
        # Enable WAL mode for better concurrency
        if self._enable_wal:
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA synchronous=NORMAL")
        
        # Set performance optimizations
        conn.execute("PRAGMA busy_timeout=5000")  # 5 second timeout
        conn.execute("PRAGMA temp_store=MEMORY")  # Store temp tables in memory
        conn.execute("PRAGMA mmap_size=268435456")  # 256MB memory-mapped I/O
        conn.execute("PRAGMA page_size=4096")  # Optimal page size
        conn.execute("PRAGMA cache_size=-64000")  # 64MB cache
        conn.execute("PRAGMA locking_mode=NORMAL")
        conn.execute("PRAGMA foreign_keys=OFF")  # No foreign keys needed
        
        return conn
    
    def _return_connection(self, conn: sqlite3.Connection) -> None:
        """Return a connection to the pool."""
        try:
            self._connection_pool.put_nowait(conn)
        except:
            # Pool is full, close the connection
            conn.close()
            with self._pool_lock:
                self._pool_created -= 1
    
    def _initialize_db(self) -> None:
        """Create database schema if it doesn't exist."""
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Create schema version table for migrations
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS schema_version (
                version INTEGER PRIMARY KEY,
                applied_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS trades (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                market_slug TEXT NOT NULL,
                market_id TEXT NOT NULL,
                side TEXT NOT NULL,
                entry_price REAL NOT NULL,
                exit_price REAL,
                amount REAL NOT NULL,
                shares REAL NOT NULL,
                fee REAL NOT NULL,
                outcome TEXT,
                pnl REAL NOT NULL,
                timestamp TEXT NOT NULL,
                market_session TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                order_id TEXT,
                status TEXT DEFAULT 'pending',
                filled_shares REAL DEFAULT 0.0,
                filled_amount REAL DEFAULT 0.0,
                avg_fill_price REAL DEFAULT 0.0,
                filled_at TEXT
            )
        """)
        
        # Create indexes for common queries
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_market_slug 
            ON trades(market_slug)
        """)
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_market_id 
            ON trades(market_id)
        """)
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_side 
            ON trades(side)
        """)
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_outcome 
            ON trades(outcome)
        """)
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_timestamp 
            ON trades(timestamp)
        """)
        
        # Create composite index for duplicate detection
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_duplicate_check 
            ON trades(market_id, side, timestamp)
        """)
        
        # Create index for market session
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_market_session 
            ON trades(market_session)
        """)
        
        # Create materialized views for common aggregations
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS trade_statistics_mv (
                asset TEXT PRIMARY KEY,
                total_trades INTEGER,
                wins INTEGER,
                losses INTEGER,
                win_rate REAL,
                total_pnl REAL,
                avg_pnl REAL,
                last_updated TEXT
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS daily_summary_mv (
                date TEXT PRIMARY KEY,
                total_trades INTEGER,
                total_pnl REAL,
                total_fees REAL,
                win_rate REAL,
                last_updated TEXT
            )
        """)
        
        # Initialize schema version if not exists
        cursor.execute("INSERT OR IGNORE INTO schema_version (version) VALUES (1)")
        
        conn.commit()
        log.info("Database initialized at %s", self.db_path)
    
    def _validate_trade_data(
        self,
        market_slug: str,
        market_id: str,
        side: str,
        entry_price: float,
        exit_price: Optional[float],
        amount: float,
        shares: float,
        fee: float,
        outcome: Optional[str],
        pnl: float,
        timestamp: datetime,
        market_session: Optional[str] = None,
    ) -> None:
        """
        Validate trade data before saving.
        
        Raises
        ------
        ValueError
            If any validation fails.
        """
        # Validate required string fields
        if not market_slug or not isinstance(market_slug, str):
            raise ValueError("market_slug must be a non-empty string")
        if not market_id or not isinstance(market_id, str):
            raise ValueError("market_id must be a non-empty string")
        
        # Validate side
        if side not in ("UP", "DOWN"):
            raise ValueError(f"side must be 'UP' or 'DOWN', got '{side}'")
        
        # Validate numeric fields
        if entry_price < 0:
            raise ValueError(f"entry_price must be non-negative, got {entry_price}")
        if exit_price is not None and exit_price < 0:
            raise ValueError(f"exit_price must be non-negative, got {exit_price}")
        if amount < 0:
            raise ValueError(f"amount must be non-negative, got {amount}")
        if shares < 0:
            raise ValueError(f"shares must be non-negative, got {shares}")
        if fee < 0:
            raise ValueError(f"fee must be non-negative, got {fee}")
        
        # Validate outcome
        valid_outcomes = {"WON", "LOST", "CLOSED", None}
        if outcome not in valid_outcomes:
            raise ValueError(f"outcome must be one of {valid_outcomes}, got '{outcome}'")
        
        # Validate timestamp
        if not isinstance(timestamp, datetime):
            raise ValueError(f"timestamp must be a datetime object, got {type(timestamp)}")
        if timestamp.tzinfo is None:
            raise ValueError("timestamp must be timezone-aware")
        
        # Validate market_session if provided
        if market_session is not None:
            valid_sessions = {"london", "new_york", "asia", "sydney"}
            if market_session not in valid_sessions:
                raise ValueError(f"market_session must be one of {valid_sessions}, got '{market_session}'")
    
    def save_trade(
        self,
        market_slug: str,
        market_id: str,
        side: str,
        entry_price: float,
        exit_price: Optional[float],
        amount: float,
        shares: float,
        fee: float,
        outcome: Optional[str],
        pnl: float,
        timestamp: datetime,
        market_session: Optional[str] = None,
        check_duplicates: bool = True,
        order_id: Optional[str] = None,
        status: str = "pending",
    ) -> int:
        """
        Save a trade to the database.
        
        Parameters
        ----------
        market_slug : str
            Market slug identifier.
        market_id : str
            Market ID from Polymarket.
        side : str
            "UP" or "DOWN".
        entry_price : float
            Entry price per share.
        exit_price : float or None
            Exit price if position was closed.
        amount : float
            USDC amount spent.
        shares : float
            Number of shares received.
        fee : float
            Fee paid in USDC.
        outcome : str or None
            "WON", "LOST", "CLOSED", or None if pending.
        pnl : float
            Profit or loss in USDC.
        timestamp : datetime
            Trade timestamp (UTC).
        market_session : str or None, optional
            Market session name (e.g., "london", "new_york", "asia", "sydney").
        check_duplicates : bool, optional
            Whether to check for duplicate trades (default: True).
        
        Returns
        -------
        int
            The ID of the inserted trade.
        
        Raises
        ------
        ValueError
            If validation fails or duplicate detected.
        """
        # Validate data
        self._validate_trade_data(
            market_slug, market_id, side, entry_price, exit_price,
            amount, shares, fee, outcome, pnl, timestamp, market_session
        )
        
        # Check for duplicates if enabled
        if check_duplicates:
            if self.is_duplicate_trade(market_id, side, timestamp):
                raise ValueError(
                    f"Duplicate trade detected: market_id={market_id}, "
                    f"side={side}, timestamp={timestamp.isoformat()}"
                )
        
        conn = self._get_connection()
        cursor = conn.cursor()
        
        timestamp_str = timestamp.isoformat()
        
        cursor.execute("""
            INSERT INTO trades (
                market_slug, market_id, side, entry_price, exit_price,
                amount, shares, fee, outcome, pnl, timestamp, market_session,
                order_id, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            market_slug, market_id, side, entry_price, exit_price,
            amount, shares, fee, outcome, pnl, timestamp_str, market_session,
            order_id, status
        ))
        
        conn.commit()
        trade_id = cursor.lastrowid
        
        # Invalidate cache on write
        self._invalidate_cache()
        
        # Trigger trade_saved hooks if streaming is enabled
        if self._streaming_enabled:
            trade_record = self._row_to_trade_record(cursor.execute(
                "SELECT * FROM trades WHERE id = ?", (trade_id,)
            ).fetchone())
            self._trigger_trade_saved_hooks(trade_record)
        
        log.debug("Trade saved: ID=%d, market=%s, side=%s, pnl=%.2f",
                  trade_id, market_slug, side, pnl)
        return trade_id
    
    def save_trades_bulk(
        self,
        trades: List[Dict[str, Any]],
        check_duplicates: bool = True,
    ) -> List[int]:
        """
        Save multiple trades in a single transaction for better performance.
        
        Parameters
        ----------
        trades : list of dict
            List of trade dictionaries with keys: market_slug, market_id, side,
            entry_price, exit_price, amount, shares, fee, outcome, pnl, timestamp, market_session.
        check_duplicates : bool, optional
            Whether to check for duplicate trades (default: True).
        
        Returns
        -------
        list of int
            List of inserted trade IDs.
        
        Raises
        ------
        ValueError
            If validation fails or duplicate detected.
        """
        if not trades:
            return []
        
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Validate all trades first
        for trade in trades:
            self._validate_trade_data(
                market_slug=trade["market_slug"],
                market_id=trade["market_id"],
                side=trade["side"],
                entry_price=trade["entry_price"],
                exit_price=trade.get("exit_price"),
                amount=trade["amount"],
                shares=trade["shares"],
                fee=trade["fee"],
                outcome=trade.get("outcome"),
                pnl=trade["pnl"],
                timestamp=trade["timestamp"],
                market_session=trade.get("market_session"),
            )
        
        # Check for duplicates if enabled
        if check_duplicates:
            for trade in trades:
                if self.is_duplicate_trade(
                    trade["market_id"],
                    trade["side"],
                    trade["timestamp"]
                ):
                    raise ValueError(
                        f"Duplicate trade detected: market_id={trade['market_id']}, "
                        f"side={trade['side']}, timestamp={trade['timestamp'].isoformat()}"
                    )
        
        # Prepare data for bulk insert
        trade_data = []
        for trade in trades:
            trade_data.append((
                trade["market_slug"],
                trade["market_id"],
                trade["side"],
                trade["entry_price"],
                trade.get("exit_price"),
                trade["amount"],
                trade["shares"],
                trade["fee"],
                trade.get("outcome"),
                trade["pnl"],
                trade["timestamp"].isoformat(),
                trade.get("market_session"),
            ))
        
        # Begin transaction
        cursor.execute("BEGIN TRANSACTION")
        
        try:
            # Bulk insert
            cursor.executemany("""
                INSERT INTO trades (
                    market_slug, market_id, side, entry_price, exit_price,
                    amount, shares, fee, outcome, pnl, timestamp, market_session
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, trade_data)
            
            conn.commit()
            
            # Get the inserted IDs by querying the last inserted rows
            # Get the last row ID
            last_id = cursor.lastrowid
            first_id = last_id - len(trades) + 1
            
            # Generate the list of IDs
            trade_ids = list(range(first_id, last_id + 1))
            
            # Invalidate cache on write
            self._invalidate_cache()
            
            log.info("Bulk saved %d trades", len(trades))
            return trade_ids

        except Exception as e:
            conn.rollback()
            log.error("Bulk insert failed: %s", e)
            raise

    def update_trade_status(
        self,
        order_id: str,
        status: str,
        filled_shares: float = 0.0,
        filled_amount: float = 0.0,
        avg_fill_price: float = 0.0,
        filled_at: Optional[datetime] = None,
    ) -> bool:
        """
        Update order fill status in the database.

        Parameters
        ----------
        order_id : str
            Order ID to update
        status : str
            New order status ("pending", "open", "filled", "partially_filled", "cancelled", "expired")
        filled_shares : float, optional
            Number of shares filled
        filled_amount : float, optional
            Amount filled in USDC
        avg_fill_price : float, optional
            Average fill price
        filled_at : datetime, optional
            Timestamp when order was filled

        Returns
        -------
        bool
            True if update was successful, False otherwise
        """
        conn = self._get_connection()
        cursor = conn.cursor()

        filled_at_str = filled_at.isoformat() if filled_at else None

        try:
            cursor.execute("""
                UPDATE trades
                SET status = ?,
                    filled_shares = ?,
                    filled_amount = ?,
                    avg_fill_price = ?,
                    filled_at = ?
                WHERE order_id = ?
            """, (status, filled_shares, filled_amount, avg_fill_price, filled_at_str, order_id))

            conn.commit()

            if cursor.rowcount > 0:
                log.debug("Trade status updated: order_id=%s, status=%s", order_id, status)
                # Invalidate cache on update
                self._invalidate_cache()
                return True
            else:
                log.warning("No trade found with order_id=%s", order_id)
                return False

        except Exception as e:
            conn.rollback()
            log.error("Failed to update trade status: %s", e)
            return False
    
    def is_duplicate_trade(
        self,
        market_id: str,
        side: str,
        timestamp: datetime,
        tolerance_seconds: int = 1,
    ) -> bool:
        """
        Check if a trade already exists in the database.
        
        A trade is considered a duplicate if it has the same market_id, side,
        and timestamp (within a tolerance window).
        
        Parameters
        ----------
        market_id : str
            Market ID to check.
        side : str
            Side to check ("UP" or "DOWN").
        timestamp : datetime
            Timestamp to check.
        tolerance_seconds : int, optional
            Time tolerance in seconds for timestamp matching (default: 1).
        
        Returns
        -------
        bool
            True if a duplicate exists, False otherwise.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        timestamp_str = timestamp.isoformat()
        
        # Check for exact match first
        cursor.execute("""
            SELECT COUNT(*) FROM trades
            WHERE market_id = ? AND side = ? AND timestamp = ?
        """, (market_id, side.upper(), timestamp_str))
        
        count = cursor.fetchone()[0]
        if count > 0:
            return True
        
        # If no exact match, check within tolerance window
        # This handles cases where timestamps might differ slightly
        cursor.execute("""
            SELECT COUNT(*) FROM trades
            WHERE market_id = ? AND side = ?
        """, (market_id, side.upper()))
        
        rows = cursor.fetchall()
        for row in rows:
            # Parse stored timestamps and check tolerance
            cursor.execute("""
                SELECT timestamp FROM trades
                WHERE market_id = ? AND side = ?
            """, (market_id, side.upper()))
            
            for ts_row in cursor.fetchall():
                stored_ts = datetime.fromisoformat(ts_row[0])
                if stored_ts.tzinfo is None:
                    stored_ts = stored_ts.replace(tzinfo=timezone.utc)
                
                time_diff = abs((timestamp - stored_ts).total_seconds())
                if time_diff <= tolerance_seconds:
                    return True
        
        return False
    
    def get_schema_version(self) -> int:
        """
        Get the current schema version.
        
        Returns
        -------
        int
            The current schema version.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT MAX(version) FROM schema_version")
        result = cursor.fetchone()
        return result[0] if result[0] is not None else 0
    
    def _apply_migration(self, version: int, migration_sql: str) -> None:
        """
        Apply a migration to the database.
        
        Parameters
        ----------
        version : int
            The migration version number.
        migration_sql : str
            SQL statements to execute for the migration.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        current_version = self.get_schema_version()
        
        if current_version >= version:
            log.debug("Migration %d already applied (current version: %d)", version, current_version)
            return
        
        log.info("Applying migration %d", version)
        
        try:
            # Execute migration SQL
            cursor.executescript(migration_sql)
            
            # Update schema version
            cursor.execute(
                "INSERT INTO schema_version (version) VALUES (?)",
                (version,)
            )
            
            conn.commit()
            log.info("Migration %d applied successfully", version)
        except Exception as e:
            conn.rollback()
            log.error("Migration %d failed: %s", version, e)
            raise
    
    def run_migrations(self) -> None:
        """
        Run all pending migrations.
        
        This method checks the current schema version and applies any
        pending migrations in order.
        """
        current_version = self.get_schema_version()
        log.info("Current schema version: %d", current_version)
        
        # Define migrations
        migrations = {
            2: """
            ALTER TABLE trades ADD COLUMN order_id TEXT;
            ALTER TABLE trades ADD COLUMN status TEXT DEFAULT 'pending';
            ALTER TABLE trades ADD COLUMN filled_shares REAL DEFAULT 0.0;
            ALTER TABLE trades ADD COLUMN filled_amount REAL DEFAULT 0.0;
            ALTER TABLE trades ADD COLUMN avg_fill_price REAL DEFAULT 0.0;
            ALTER TABLE trades ADD COLUMN filled_at TEXT;
            """,
        }
        
        # Apply migrations in order
        for version, migration_sql in sorted(migrations.items()):
            if version > current_version:
                self._apply_migration(version, migration_sql)
    
    def _generate_cache_key(
        self,
        filters: Optional[Dict[str, Any]] = None,
        sort_by: str = "timestamp",
        sort_order: str = "desc",
        limit: Optional[int] = None,
        offset: int = 0,
    ) -> str:
        """Generate a cache key for query parameters."""
        key_data = {
            "filters": filters,
            "sort_by": sort_by,
            "sort_order": sort_order,
            "limit": limit,
            "offset": offset,
        }
        key_str = json.dumps(key_data, sort_keys=True)
        return hashlib.md5(key_str.encode()).hexdigest()
    
    def _invalidate_cache(self) -> None:
        """Invalidate the query cache."""
        self._query_cache.clear()
        self._cache_ttl.clear()
        log.debug("Query cache invalidated")
    
    def _is_cache_entry_valid(self, cache_key: str) -> bool:
        """Check if a cache entry is still valid based on TTL."""
        if cache_key not in self._cache_ttl:
            return False
        
        entry_time = self._cache_ttl[cache_key]
        current_time = time.time()
        return (current_time - entry_time) < self._default_cache_ttl
    
    def _get_prepared_statement(self, query: str) -> sqlite3.Statement:
        """Get or create a prepared statement for the given query."""
        with self._statement_lock:
            if query in self._prepared_statements:
                return self._prepared_statements[query]
            
            conn = self._get_connection()
            stmt = conn.prepare(query)
            
            # Implement LRU eviction for statement cache
            if len(self._prepared_statements) >= self._statement_cache_max_size:
                # Remove oldest entry
                self._prepared_statements.pop(next(iter(self._prepared_statements)))
            
            self._prepared_statements[query] = stmt
            log.debug("Prepared statement cached for query: %s", query[:50])
            return stmt
    
    def clear_prepared_statements(self) -> None:
        """Clear the prepared statement cache."""
        with self._statement_lock:
            self._prepared_statements.clear()
            log.debug("Prepared statement cache cleared")
    
    def refresh_materialized_views(self) -> None:
        """Refresh materialized views for common aggregations."""
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Refresh trade statistics by asset
        cursor.execute("DELETE FROM trade_statistics_mv")
        
        cursor.execute("""
            INSERT INTO trade_statistics_mv (asset, total_trades, wins, losses, win_rate, total_pnl, avg_pnl, last_updated)
            SELECT 
                SUBSTR(market_slug, 1, INSTR(market_slug, '-') - 1) as asset,
                COUNT(*) as total_trades,
                SUM(CASE WHEN outcome = 'WON' THEN 1 ELSE 0 END) as wins,
                SUM(CASE WHEN outcome = 'LOST' THEN 1 ELSE 0 END) as losses,
                CAST(SUM(CASE WHEN outcome = 'WON' THEN 1 ELSE 0 END) * 100.0 / COUNT(*) AS REAL) as win_rate,
                SUM(pnl) as total_pnl,
                AVG(pnl) as avg_pnl,
                datetime('now') as last_updated
            FROM trades
            GROUP BY asset
        """)
        
        # Refresh daily summary
        cursor.execute("DELETE FROM daily_summary_mv")
        
        cursor.execute("""
            INSERT INTO daily_summary_mv (date, total_trades, total_pnl, total_fees, win_rate, last_updated)
            SELECT 
                DATE(timestamp) as date,
                COUNT(*) as total_trades,
                SUM(pnl) as total_pnl,
                SUM(fee) as total_fees,
                CAST(SUM(CASE WHEN outcome = 'WON' THEN 1 ELSE 0 END) * 100.0 / COUNT(*) AS REAL) as win_rate,
                datetime('now') as last_updated
            FROM trades
            GROUP BY DATE(timestamp)
        """)
        
        conn.commit()
        log.info("Materialized views refreshed")
    
    def get_trade_statistics_from_mv(self) -> Dict[str, Dict[str, Any]]:
        """Get trade statistics from materialized view (faster than calculating)."""
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM trade_statistics_mv")
        
        stats = {}
        for row in cursor.fetchall():
            stats[row['asset']] = {
                'total_trades': row['total_trades'],
                'wins': row['wins'],
                'losses': row['losses'],
                'win_rate': row['win_rate'],
                'total_pnl': row['total_pnl'],
                'avg_pnl': row['avg_pnl'],
                'last_updated': row['last_updated']
            }
        
        return stats
    
    def get_daily_summary_from_mv(self) -> Dict[str, Dict[str, Any]]:
        """Get daily summary from materialized view."""
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM daily_summary_mv ORDER BY date DESC")
        
        summary = {}
        for row in cursor.fetchall():
            summary[row['date']] = {
                'total_trades': row['total_trades'],
                'total_pnl': row['total_pnl'],
                'total_fees': row['total_fees'],
                'win_rate': row['win_rate'],
                'last_updated': row['last_updated']
            }
        
        return summary
    
    def start_background_optimization(self, interval_seconds: int = 3600) -> None:
        """
        Start background optimization thread.
        
        Parameters
        ----------
        interval_seconds : int, optional
            Interval between optimizations in seconds (default: 3600 = 1 hour).
        """
        if self._optimization_thread is not None and self._optimization_thread.is_alive():
            log.warning("Background optimization thread already running")
            return
        
        self._optimization_interval = interval_seconds
        self._optimization_stop_event.clear()
        
        def optimization_loop():
            while not self._optimization_stop_event.wait(self._optimization_interval):
                try:
                    self._run_background_optimization()
                except Exception as e:
                    log.error("Background optimization failed: %s", e)
        
        self._optimization_thread = Thread(target=optimization_loop, daemon=True)
        self._optimization_thread.start()
        log.info("Background optimization started (interval: %d seconds)", interval_seconds)
    
    def stop_background_optimization(self) -> None:
        """Stop background optimization thread."""
        if self._optimization_thread is not None:
            self._optimization_stop_event.set()
            self._optimization_thread.join(timeout=5)
            self._optimization_thread = None
            log.info("Background optimization stopped")
    
    def _run_background_optimization(self) -> None:
        """Run background optimization tasks."""
        log.info("Running background optimization...")
        
        # Refresh materialized views
        self.refresh_materialized_views()
        
        # Analyze indexes
        self.analyze_indexes()
        
        # Optimize database
        self.optimize_database()
        
        # Clean expired cache entries
        self._clean_expired_cache()
        
        log.info("Background optimization completed")
    
    def _clean_expired_cache(self) -> None:
        """Remove expired cache entries based on TTL."""
        current_time = time.time()
        expired_keys = []
        
        for cache_key, entry_time in self._cache_ttl.items():
            if (current_time - entry_time) >= self._default_cache_ttl:
                expired_keys.append(cache_key)
        
        for key in expired_keys:
            self._query_cache.pop(key, None)
            self._cache_ttl.pop(key, None)
        
        if expired_keys:
            log.debug("Cleaned %d expired cache entries", len(expired_keys))
    
    def stream_trades(
        self,
        filters: Optional[Dict[str, Any]] = None,
        batch_size: int = 100,
    ) -> List[TradeRecord]:
        """
        Stream trades in batches for large datasets.
        
        This is a generator that yields batches of trades, useful for processing
        large datasets without loading everything into memory.
        
        Parameters
        ----------
        filters : dict, optional
            Filter criteria (same as load_trades()).
        batch_size : int, optional
            Number of trades per batch (default: 100).
        
        Yields
        ------
        List[TradeRecord]
            Batch of trades.
        
        Example
        -------
        >>> for batch in db.stream_trades(batch_size=500):
        ...     process_batch(batch)
        """
        offset = 0
        while True:
            batch = self.load_trades(filters=filters, limit=batch_size, offset=offset)
            if not batch:
                break
            yield batch
            offset += batch_size
    
    def stream_trades_by_asset(self, asset: str, batch_size: int = 100) -> List[TradeRecord]:
        """
        Stream trades for a specific asset in batches.
        
        Parameters
        ----------
        asset : str
            Asset symbol to filter by.
        batch_size : int, optional
            Number of trades per batch (default: 100).
        
        Yields
        ------
        List[TradeRecord]
            Batch of trades for the asset.
        """
        offset = 0
        while True:
            conn = self._get_connection()
            cursor = conn.cursor()
            
            pattern = f"{asset.lower()}%"
            cursor.execute("""
                SELECT id, market_slug, market_id, side, entry_price, exit_price,
                       amount, shares, fee, outcome, pnl, timestamp, market_session
                FROM trades
                WHERE LOWER(market_slug) LIKE ?
                ORDER BY timestamp DESC
                LIMIT ? OFFSET ?
            """, (pattern, batch_size, offset))
            
            batch = []
            for row in cursor.fetchall():
                batch.append(self._row_to_trade_record(row))
            
            if not batch:
                break
            
            yield batch
            offset += batch_size
    
    def enable_cache(self) -> None:
        """Enable query result caching."""
        self._cache_enabled = True
        log.debug("Query cache enabled")
    
    def disable_cache(self) -> None:
        """Disable query result caching and clear existing cache."""
        self._cache_enabled = False
        self._invalidate_cache()
        log.debug("Query cache disabled")
    
    def execute_parallel_queries(
        self,
        queries: List[str],
        params_list: Optional[List[tuple]] = None,
        max_workers: int = 4,
    ) -> List[List[sqlite3.Row]]:
        """
        Execute multiple queries in parallel using thread pool.
        
        This is useful for running independent queries concurrently,
        such as fetching statistics for multiple assets simultaneously.
        
        Parameters
        ----------
        queries : list of str
            List of SQL queries to execute.
        params_list : list of tuple, optional
            List of parameter tuples for each query.
        max_workers : int, optional
            Maximum number of worker threads (default: 4).
        
        Returns
        -------
        list of list of sqlite3.Row
            Results for each query in the same order as input.
        
        Example
        -------
        >>> queries = [
        ...     "SELECT * FROM trades WHERE side = 'UP'",
        ...     "SELECT * FROM trades WHERE side = 'DOWN'",
        ...     "SELECT * FROM trades WHERE outcome = 'WON'"
        ... ]
        >>> results = db.execute_parallel_queries(queries)
        """
        if params_list is None:
            params_list = [() for _ in queries]
        
        if len(queries) != len(params_list):
            raise ValueError("Number of queries must match number of params lists")
        
        results = [None] * len(queries)
        
        def execute_query(index: int, query: str, params: tuple) -> None:
            """Execute a single query and store result."""
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute(query, params)
            results[index] = cursor.fetchall()
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [
                executor.submit(execute_query, i, q, p)
                for i, (q, p) in enumerate(zip(queries, params_list))
            ]
            
            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    log.error("Parallel query failed: %s", e)
        
        return results
    
    def get_parallel_statistics_by_assets(
        self,
        assets: List[str],
        max_workers: int = 4,
    ) -> Dict[str, Dict[str, Any]]:
        """
        Get statistics for multiple assets in parallel.
        
        Parameters
        ----------
        assets : list of str
            List of asset symbols (e.g., ["BTC", "ETH", "SOL"]).
        max_workers : int, optional
            Maximum number of worker threads (default: 4).
        
        Returns
        -------
        dict
            Dictionary mapping asset names to their statistics.
        
        Example
        -------
        >>> stats = db.get_parallel_statistics_by_assets(["BTC", "ETH", "SOL"])
        >>> print(stats["BTC"]["total_trades"])
        """
        queries = []
        params_list = []
        
        for asset in assets:
            pattern = f"{asset.lower()}%"
            queries.append("""
                SELECT 
                    COUNT(*) as total_trades,
                    SUM(CASE WHEN outcome = 'WON' THEN 1 ELSE 0 END) as wins,
                    SUM(CASE WHEN outcome = 'LOST' THEN 1 ELSE 0 END) as losses,
                    SUM(pnl) as total_pnl,
                    AVG(pnl) as avg_pnl,
                    SUM(fee) as total_fees
                FROM trades
                WHERE LOWER(market_slug) LIKE ?
            """)
            params_list.append((pattern,))
        
        results = self.execute_parallel_queries(queries, params_list, max_workers)
        
        stats = {}
        for asset, rows in zip(assets, results):
            if rows and rows[0]:
                row = rows[0]
                total_trades = row['total_trades']
                wins = row['wins'] or 0
                losses = row['losses'] or 0
                total_pnl = row['total_pnl'] or 0
                avg_pnl = row['avg_pnl'] or 0
                total_fees = row['total_fees'] or 0
                win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
                
                stats[asset] = {
                    'total_trades': total_trades,
                    'wins': wins,
                    'losses': losses,
                    'win_rate': win_rate,
                    'total_pnl': total_pnl,
                    'avg_pnl': avg_pnl,
                    'total_fees': total_fees,
                }
            else:
                stats[asset] = {
                    'total_trades': 0,
                    'wins': 0,
                    'losses': 0,
                    'win_rate': 0,
                    'total_pnl': 0,
                    'avg_pnl': 0,
                    'total_fees': 0,
                }
        
        return stats
    
    def clear_cache(self) -> None:
        """Clear the query cache."""
        self._invalidate_cache()
    
    def analyze_indexes(self) -> None:
        """
        Analyze database indexes to optimize query planning.
        
        This runs SQLite's ANALYZE command to update statistics
        that help the query planner choose optimal execution plans.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("ANALYZE")
        conn.commit()
        
        log.info("Database indexes analyzed")
    
    def optimize_database(self) -> None:
        """
        Optimize the database for better performance.
        
        This runs SQLite's OPTIMIZE command which can:
        - Rebuild the database file
        - Update statistics
        - Defragment the database
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("PRAGMA optimize")
        conn.commit()
        
        self._last_optimization = datetime.now(timezone.utc)
        log.info("Database optimized")
    
    def get_index_info(self) -> Dict[str, Dict[str, Any]]:
        """
        Get information about database indexes.
        
        Returns
        -------
        dict
            Dictionary with index names as keys and index details as values.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT name, tbl_name, sql 
            FROM sqlite_master 
            WHERE type='index' AND name NOT LIKE 'sqlite_%'
        """)
        
        indexes = {}
        for row in cursor.fetchall():
            indexes[row["name"]] = {
                "table": row["tbl_name"],
                "sql": row["sql"],
            }
        
        return indexes
    
    def rebuild_index(self, index_name: str) -> None:
        """
        Rebuild a specific index to optimize performance.
        
        Parameters
        ----------
        index_name : str
            Name of the index to rebuild.
        
        Raises
        ------
        ValueError
            If the index doesn't exist.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Validate index name using whitelist to prevent SQL injection
        valid_indexes = {
            "idx_market_slug": "CREATE INDEX idx_market_slug ON trades(market_slug)",
            "idx_market_id": "CREATE INDEX idx_market_id ON trades(market_id)",
            "idx_side": "CREATE INDEX idx_side ON trades(side)",
            "idx_outcome": "CREATE INDEX idx_outcome ON trades(outcome)",
            "idx_timestamp": "CREATE INDEX idx_timestamp ON trades(timestamp)",
            "idx_duplicate_check": "CREATE INDEX idx_duplicate_check ON trades(market_id, side, timestamp)",
            "idx_market_session": "CREATE INDEX idx_market_session ON trades(market_session)"
        }
        
        if index_name not in valid_indexes:
            raise ValueError(f"Invalid index name '{index_name}'. Valid options: {sorted(valid_indexes.keys())}")
        
        # Check if index exists
        cursor.execute("""
            SELECT name FROM sqlite_master 
            WHERE type='index' AND name=?
        """, (index_name,))
        
        if not cursor.fetchone():
            raise ValueError(f"Index '{index_name}' does not exist")
        
        # Rebuild by dropping and recreating using validated index name
        # SQLite doesn't support parameterized identifiers, but whitelist validation ensures safety
        cursor.execute(f"DROP INDEX IF EXISTS {index_name}")
        
        # Recreate based on the index type
        cursor.execute(valid_indexes[index_name])
        
        conn.commit()
        log.info("Index '%s' rebuilt", index_name)
    
    def load_all_trades(self) -> List[TradeRecord]:
        """
        Load all trades from the database.
        
        Returns
        -------
        List[TradeRecord]
            All trades in the database.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, market_slug, market_id, side, entry_price, exit_price,
                   amount, shares, fee, outcome, pnl, timestamp, market_session
            FROM trades
            ORDER BY timestamp DESC
        """)
        
        trades = []
        for row in cursor.fetchall():
            trades.append(self._row_to_trade_record(row))
        
        log.debug("Loaded %d trades from database", len(trades))
        return trades
    
    def load_trades_by_market(self, market_slug: str) -> List[TradeRecord]:
        """
        Load trades for a specific market slug.
        
        Parameters
        ----------
        market_slug : str
            Market slug to filter by.
        
        Returns
        -------
        List[TradeRecord]
            Trades for the specified market.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, market_slug, market_id, side, entry_price, exit_price,
                   amount, shares, fee, outcome, pnl, timestamp, market_session
            FROM trades
            WHERE market_slug = ?
            ORDER BY timestamp DESC
        """, (market_slug,))
        
        trades = []
        for row in cursor.fetchall():
            trades.append(self._row_to_trade_record(row))
        
        return trades
    
    def load_trades_by_asset(self, asset: str) -> List[TradeRecord]:
        """
        Load trades for a specific asset (e.g., "BTC", "ETH").
        
        Parameters
        ----------
        asset : str
            Asset symbol to filter by (case-insensitive).
        
        Returns
        -------
        List[TradeRecord]
            Trades for the specified asset.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Match market_slug starting with asset (case-insensitive)
        pattern = f"{asset.lower()}%"
        cursor.execute("""
            SELECT id, market_slug, market_id, side, entry_price, exit_price,
                   amount, shares, fee, outcome, pnl, timestamp, market_session
            FROM trades
            WHERE LOWER(market_slug) LIKE ?
            ORDER BY timestamp DESC
        """, (pattern,))
        
        trades = []
        for row in cursor.fetchall():
            trades.append(self._row_to_trade_record(row))
        
        return trades
    
    def load_trades_by_side(self, side: str) -> List[TradeRecord]:
        """
        Load trades for a specific side.
        
        Parameters
        ----------
        side : str
            "UP" or "DOWN".
        
        Returns
        -------
        List[TradeRecord]
            Trades for the specified side.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, market_slug, market_id, side, entry_price, exit_price,
                   amount, shares, fee, outcome, pnl, timestamp, market_session
            FROM trades
            WHERE side = ?
            ORDER BY timestamp DESC
        """, (side.upper(),))
        
        trades = []
        for row in cursor.fetchall():
            trades.append(self._row_to_trade_record(row))
        
        return trades
    
    def load_trades_by_outcome(self, outcome: str) -> List[TradeRecord]:
        """
        Load trades with a specific outcome.
        
        Parameters
        ----------
        outcome : str
            "WON", "LOST", or "CLOSED".
        
        Returns
        -------
        List[TradeRecord]
            Trades with the specified outcome.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, market_slug, market_id, side, entry_price, exit_price,
                   amount, shares, fee, outcome, pnl, timestamp, market_session
            FROM trades
            WHERE outcome = ?
            ORDER BY timestamp DESC
        """, (outcome.upper(),))
        
        trades = []
        for row in cursor.fetchall():
            trades.append(self._row_to_trade_record(row))
        
        return trades
    
    def load_trades_by_market_session(self, market_session: str) -> List[TradeRecord]:
        """
        Load trades for a specific market session.
        
        Parameters
        ----------
        market_session : str
            Market session name ("london", "new_york", "asia", "sydney").
        
        Returns
        -------
        List[TradeRecord]
            Trades for the specified market session.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, market_slug, market_id, side, entry_price, exit_price,
                   amount, shares, fee, outcome, pnl, timestamp, market_session
            FROM trades
            WHERE market_session = ?
            ORDER BY timestamp DESC
        """, (market_session,))
        
        trades = []
        for row in cursor.fetchall():
            trades.append(self._row_to_trade_record(row))
        
        return trades
    
    def load_trades_by_date_range(
        self,
        start_date: datetime,
        end_date: datetime
    ) -> List[TradeRecord]:
        """
        Load trades within a date range.
        
        Parameters
        ----------
        start_date : datetime
            Start of date range (inclusive).
        end_date : datetime
            End of date range (inclusive).
        
        Returns
        -------
        List[TradeRecord]
            Trades within the specified date range.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, market_slug, market_id, side, entry_price, exit_price,
                   amount, shares, fee, outcome, pnl, timestamp, market_session
            FROM trades
            WHERE timestamp >= ? AND timestamp <= ?
            ORDER BY timestamp DESC
        """, (start_date.isoformat(), end_date.isoformat()))
        
        trades = []
        for row in cursor.fetchall():
            trades.append(self._row_to_trade_record(row))
        
        return trades
    
    def load_trades(
        self,
        filters: Optional[Dict[str, Any]] = None,
        sort_by: str = "timestamp",
        sort_order: str = "desc",
        limit: Optional[int] = None,
        offset: int = 0,
    ) -> List[TradeRecord]:
        """
        Load trades with advanced filtering, sorting, and pagination.
        
        Parameters
        ----------
        filters : dict, optional
            Dictionary of filter criteria. Supported keys:
            - asset: str (e.g., "BTC")
            - side: str ("UP" or "DOWN")
            - outcome: str ("WON", "LOST", "CLOSED")
            - min_pnl: float (minimum P&L)
            - max_pnl: float (maximum P&L)
            - min_amount: float (minimum amount)
            - max_amount: float (maximum amount)
            - market_slug: str (exact match)
            - market_id: str (exact match)
        sort_by : str, optional
            Field to sort by (default: "timestamp").
            Options: "timestamp", "pnl", "amount", "entry_price", "shares", "fee"
        sort_order : str, optional
            Sort order: "asc" or "desc" (default: "desc").
        limit : int, optional
            Maximum number of trades to return (default: None = no limit).
        offset : int, optional
            Number of trades to skip (default: 0).
        
        Returns
        -------
        List[TradeRecord]
            Filtered and sorted trades.
        
        Example
        -------
        >>> trades = db.load_trades(
        ...     filters={"asset": "BTC", "side": "UP", "outcome": "WON"},
        ...     sort_by="pnl",
        ...     sort_order="desc",
        ...     limit=10
        ... )
        """
        # Check cache if enabled
        if self._cache_enabled:
            cache_key = self._generate_cache_key(filters, sort_by, sort_order, limit, offset)
            if cache_key in self._query_cache and self._is_cache_entry_valid(cache_key):
                self._cache_hits += 1
                log.debug("Cache hit for key: %s", cache_key)
                return self._query_cache[cache_key]
            else:
                self._cache_misses += 1
        
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Build query with filters
        query = """
            SELECT id, market_slug, market_id, side, entry_price, exit_price,
                   amount, shares, fee, outcome, pnl, timestamp, market_session
            FROM trades
        """
        
        params = []
        where_clauses = []
        
        if filters:
            # Asset filter (pattern match on market_slug)
            if "asset" in filters:
                pattern = f"{filters['asset'].lower()}%"
                where_clauses.append("LOWER(market_slug) LIKE ?")
                params.append(pattern)
            
            # Side filter
            if "side" in filters:
                where_clauses.append("side = ?")
                params.append(filters["side"].upper())
            
            # Outcome filter
            if "outcome" in filters:
                where_clauses.append("outcome = ?")
                params.append(filters["outcome"].upper())
            
            # P&L range filters
            if "min_pnl" in filters:
                where_clauses.append("pnl >= ?")
                params.append(float(filters["min_pnl"]))
            
            if "max_pnl" in filters:
                where_clauses.append("pnl <= ?")
                params.append(float(filters["max_pnl"]))
            
            # Amount range filters
            if "min_amount" in filters:
                where_clauses.append("amount >= ?")
                params.append(float(filters["min_amount"]))
            
            if "max_amount" in filters:
                where_clauses.append("amount <= ?")
                params.append(float(filters["max_amount"]))
            
            # Exact market_slug filter
            if "market_slug" in filters:
                where_clauses.append("market_slug = ?")
                params.append(filters["market_slug"])
            
            # Exact market_id filter
            if "market_id" in filters:
                where_clauses.append("market_id = ?")
                params.append(filters["market_id"])
        
        # Add WHERE clause if filters exist
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
        
        # Validate sort_by field
        valid_sort_fields = {
            "timestamp", "pnl", "amount", "entry_price", "shares", "fee",
            "market_slug", "side", "outcome"
        }
        if sort_by not in valid_sort_fields:
            raise ValueError(
                f"Invalid sort_by field '{sort_by}'. "
                f"Valid options: {sorted(valid_sort_fields)}"
            )
        
        # Validate sort_order
        sort_order = sort_order.lower()
        if sort_order not in ("asc", "desc"):
            raise ValueError(f"sort_order must be 'asc' or 'desc', got '{sort_order}'")
        
        # Add ORDER BY clause using whitelist to prevent SQL injection
        sort_field_map = {
            "timestamp": "timestamp",
            "pnl": "pnl",
            "amount": "amount",
            "entry_price": "entry_price",
            "shares": "shares",
            "fee": "fee",
            "market_slug": "market_slug",
            "side": "side",
            "outcome": "outcome"
        }
        safe_sort_by = sort_field_map.get(sort_by)
        if not safe_sort_by:
            raise ValueError(
                f"Invalid sort_by field '{sort_by}'. "
                f"Valid options: {sorted(valid_sort_fields)}"
            )
        query += f" ORDER BY {safe_sort_by} {sort_order.upper()}"
        
        # Add LIMIT and OFFSET for pagination
        if limit is not None:
            if limit <= 0:
                raise ValueError(f"limit must be positive, got {limit}")
            query += " LIMIT ?"
            params.append(int(limit))
        
        if offset > 0:
            query += " OFFSET ?"
            params.append(int(offset))
        
        # Execute query
        cursor.execute(query, params)
        
        # Convert rows to TradeRecord objects
        trades = []
        for row in cursor.fetchall():
            trades.append(self._row_to_trade_record(row))
        
        log.debug(
            "Loaded %d trades with filters=%s, sort_by=%s, limit=%s, offset=%d",
            len(trades), filters, sort_by, limit, offset
        )
        
        # Store in cache if enabled
        if self._cache_enabled:
            cache_key = self._generate_cache_key(filters, sort_by, sort_order, limit, offset)
            # Implement LRU eviction if cache is full
            if len(self._query_cache) >= self._cache_max_size:
                # Remove oldest entry (first in dict)
                oldest_key = next(iter(self._query_cache))
                self._query_cache.pop(oldest_key)
                self._cache_ttl.pop(oldest_key, None)
            self._query_cache[cache_key] = trades
            self._cache_ttl[cache_key] = time.time()
            log.debug("Cached %d trades with key: %s", len(trades), cache_key)
        
        return trades
    
    def aggregate_trades(
        self,
        group_by: str = "asset",
        filters: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Dict[str, Any]]:
        """
        Aggregate trades by a specified field.
        
        Parameters
        ----------
        group_by : str
            Field to group by. Options: "asset", "side", "outcome", "market_slug".
        filters : dict, optional
            Same filter criteria as load_trades().
        
        Returns
        -------
        dict
            Dictionary with group keys as keys and statistics as values.
            Each value contains: count, total_pnl, avg_pnl, wins, losses, win_rate.
        
        Example
        -------
        >>> by_asset = db.aggregate_trades(group_by="asset")
        >>> print(by_asset["BTC"])
        >>> {'count': 10, 'total_pnl': 50.0, 'avg_pnl': 5.0, 'wins': 6, 'losses': 4, 'win_rate': 60.0}
        """
        # Validate group_by field
        valid_group_fields = {"asset", "side", "outcome", "market_slug"}
        if group_by not in valid_group_fields:
            raise ValueError(
                f"Invalid group_by field '{group_by}'. "
                f"Valid options: {sorted(valid_group_fields)}"
            )
        
        # Load trades with filters
        trades = self.load_trades(filters=filters)
        
        # Group trades
        groups: Dict[str, List[TradeRecord]] = {}
        
        for trade in trades:
            if group_by == "asset":
                # Extract asset from market_slug (first part before dash)
                key = trade.market_slug.split("-")[0].upper()
            elif group_by == "side":
                key = trade.side
            elif group_by == "outcome":
                key = trade.outcome or "PENDING"
            elif group_by == "market_slug":
                key = trade.market_slug
            else:
                key = str(getattr(trade, group_by, "unknown"))
            
            if key not in groups:
                groups[key] = []
            groups[key].append(trade)
        
        # Calculate statistics for each group
        results = {}
        for key, group_trades in groups.items():
            count = len(group_trades)
            total_pnl = sum(t.pnl for t in group_trades)
            avg_pnl = total_pnl / count if count > 0 else 0.0
            wins = sum(1 for t in group_trades if t.outcome == "WON")
            losses = sum(1 for t in group_trades if t.outcome == "LOST")
            win_rate = (wins / count * 100) if count > 0 else 0.0
            
            results[key] = {
                "count": count,
                "total_pnl": total_pnl,
                "avg_pnl": avg_pnl,
                "wins": wins,
                "losses": losses,
                "win_rate": win_rate,
            }
        
        log.debug("Aggregated %d trades by %s into %d groups", len(trades), group_by, len(results))
        return results
    
    def get_statistics(self) -> TradeStatistics:
        """
        Calculate statistics for all trades in the database.
        
        Returns
        -------
        TradeStatistics
            Summary statistics.
        """
        trades = self.load_all_trades()
        
        total_trades = len(trades)
        wins = sum(1 for t in trades if t.outcome == "WON")
        losses = sum(1 for t in trades if t.outcome == "LOST")
        win_rate = (wins / total_trades * 100) if total_trades > 0 else 0.0
        total_pnl = sum(t.pnl for t in trades)
        total_fees = sum(t.fee for t in trades)
        avg_entry_price = (sum(t.entry_price for t in trades) / total_trades) if total_trades > 0 else 0.0
        avg_pnl_per_trade = (total_pnl / total_trades) if total_trades > 0 else 0.0
        
        return TradeStatistics(
            total_trades=total_trades,
            wins=wins,
            losses=losses,
            win_rate=win_rate,
            total_pnl=total_pnl,
            total_fees=total_fees,
            avg_entry_price=avg_entry_price,
            avg_pnl_per_trade=avg_pnl_per_trade,
        )
    
    def export_csv(self, filepath: str | Path, filters: Optional[Dict[str, Any]] = None) -> None:
        """
        Export trades to CSV format.
        
        Parameters
        ----------
        filepath : str or Path
            Path to the output CSV file.
        filters : dict, optional
            Filter criteria to apply before export (same as load_trades()).
        
        Example
        -------
        >>> db.export_csv("trades.csv")
        >>> db.export_csv("btc_trades.csv", filters={"asset": "BTC"})
        """
        filepath = Path(filepath)
        trades = self.load_trades(filters=filters)
        
        if not trades:
            log.warning("No trades to export to CSV")
            return
        
        # Define CSV columns
        fieldnames = [
            "id", "market_slug", "market_id", "side", "entry_price",
            "exit_price", "amount", "shares", "fee", "outcome", "pnl", "timestamp"
        ]
        
        with filepath.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for trade in trades:
                row = {
                    "id": trade.id,
                    "market_slug": trade.market_slug,
                    "market_id": trade.market_id,
                    "side": trade.side,
                    "entry_price": trade.entry_price,
                    "exit_price": trade.exit_price,
                    "amount": trade.amount,
                    "shares": trade.shares,
                    "fee": trade.fee,
                    "outcome": trade.outcome,
                    "pnl": trade.pnl,
                    "timestamp": trade.timestamp.isoformat(),
                }
                writer.writerow(row)
        
        log.info("Exported %d trades to CSV: %s", len(trades), filepath)
    
    def export_json(self, filepath: str | Path, filters: Optional[Dict[str, Any]] = None) -> None:
        """
        Export trades to JSON format with metadata.
        
        Parameters
        ----------
        filepath : str or Path
            Path to the output JSON file.
        filters : dict, optional
            Filter criteria to apply before export (same as load_trades()).
        
        Example
        -------
        >>> db.export_json("trades.json")
        >>> db.export_json("won_trades.json", filters={"outcome": "WON"})
        """
        filepath = Path(filepath)
        trades = self.load_trades(filters=filters)
        
        if not trades:
            log.warning("No trades to export to JSON")
            return
        
        # Create export with metadata
        export_data = {
            "metadata": {
                "export_timestamp": datetime.now(timezone.utc).isoformat(),
                "total_trades": len(trades),
                "database_path": str(self.db_path),
            },
            "trades": [trade.to_dict() for trade in trades]
        }
        
        with filepath.open("w", encoding="utf-8") as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        log.info("Exported %d trades to JSON: %s", len(trades), filepath)
    
    def export_parquet(self, filepath: str | Path, filters: Optional[Dict[str, Any]] = None) -> None:
        """
        Export trades to Parquet format for data science workflows.
        
        This method requires the 'pyarrow' library to be installed.
        
        Parameters
        ----------
        filepath : str or Path
            Path to the output Parquet file.
        filters : dict, optional
            Filter criteria to apply before export (same as load_trades()).
        
        Raises
        ------
        ImportError
            If pyarrow is not installed.
        
        Example
        -------
        >>> db.export_parquet("trades.parquet")
        >>> db.export_parquet("btc_trades.parquet", filters={"asset": "BTC"})
        """
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError as e:
            raise ImportError(
                "pyarrow is required for Parquet export. "
                "Install it with: pip install pyarrow"
            ) from e
        
        filepath = Path(filepath)
        trades = self.load_trades(filters=filters)
        
        if not trades:
            log.warning("No trades to export to Parquet")
            return
        
        # Convert trades to list of dictionaries
        data = [trade.to_dict() for trade in trades]
        
        # Create PyArrow table
        table = pa.Table.from_pylist(data)
        
        # Write to Parquet
        pq.write_table(table, filepath)
        
        log.info("Exported %d trades to Parquet: %s", len(trades), filepath)
    
    def export_excel(self, filepath: str | Path, filters: Optional[Dict[str, Any]] = None) -> None:
        """
        Export trades to Excel format for business users.
        
        This method requires the 'openpyxl' library to be installed.
        
        Parameters
        ----------
        filepath : str or Path
            Path to the output Excel file (.xlsx).
        filters : dict, optional
            Filter criteria to apply before export (same as load_trades()).
        
        Raises
        ------
        ImportError
            If openpyxl is not installed.
        
        Example
        -------
        >>> db.export_excel("trades.xlsx")
        >>> db.export_excel("btc_trades.xlsx", filters={"asset": "BTC"})
        """
        try:
            from openpyxl import Workbook
        except ImportError as e:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            ) from e
        
        filepath = Path(filepath)
        trades = self.load_trades(filters=filters)
        
        if not trades:
            log.warning("No trades to export to Excel")
            return
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        
        # Define headers
        headers = [
            "ID", "Market Slug", "Market ID", "Side", "Entry Price",
            "Exit Price", "Amount", "Shares", "Fee", "Outcome", "P&L", "Timestamp"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Write trade data
        for row_num, trade in enumerate(trades, 2):
            ws.cell(row=row_num, column=1, value=trade.id)
            ws.cell(row=row_num, column=2, value=trade.market_slug)
            ws.cell(row=row_num, column=3, value=trade.market_id)
            ws.cell(row=row_num, column=4, value=trade.side)
            ws.cell(row=row_num, column=5, value=trade.entry_price)
            ws.cell(row=row_num, column=6, value=trade.exit_price)
            ws.cell(row=row_num, column=7, value=trade.amount)
            ws.cell(row=row_num, column=8, value=trade.shares)
            ws.cell(row=row_num, column=9, value=trade.fee)
            ws.cell(row=row_num, column=10, value=trade.outcome)
            ws.cell(row=row_num, column=11, value=trade.pnl)
            ws.cell(row=row_num, column=12, value=trade.timestamp.isoformat())
        
        # Save workbook
        wb.save(filepath)
        
        log.info("Exported %d trades to Excel: %s", len(trades), filepath)
    
    # Backup and Restore
    
    def backup(self, backup_path: str | Path) -> None:
        """
        Create a backup of the database to a local file.
        
        This method creates a full backup of the SQLite database by copying
        the database file to the specified backup location. The backup includes
        all tables, indexes, and data.
        
        Parameters
        ----------
        backup_path : str or Path
            Path where the backup file should be saved. The parent directory
            will be created if it doesn't exist.
        
        Raises
        ------
        FileNotFoundError
            If the source database file doesn't exist.
        IOError
            If the backup cannot be created due to permission or disk space issues.
        
        Example
        -------
        >>> db.backup("backups/trades_backup_2024_01_01.db")
        >>> db.backup("backups/trades_backup.db")
        """
        backup_path = Path(backup_path)
        
        # Ensure source database exists
        if not self.db_path.exists():
            raise FileNotFoundError(f"Source database not found: {self.db_path}")
        
        # Create parent directory if it doesn't exist
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Close connection before copying to ensure data integrity
        was_open = self._conn is not None
        if was_open:
            self.close()
        
        try:
            # Copy the database file
            shutil.copy2(self.db_path, backup_path)
            log.info("Database backup created: %s -> %s", self.db_path, backup_path)
        finally:
            # Reopen connection if it was open
            if was_open:
                self._get_connection()
    
    def restore(self, backup_path: str | Path, overwrite: bool = False) -> None:
        """
        Restore the database from a local backup file.
        
        This method restores the database from a previously created backup file.
        By default, this will not overwrite an existing database unless overwrite=True.
        
        Parameters
        ----------
        backup_path : str or Path
            Path to the backup file to restore from.
        overwrite : bool, optional
            If True, overwrite the existing database file. If False (default),
            raise an error if the database file already exists.
        
        Raises
        ------
        FileNotFoundError
            If the backup file doesn't exist.
        FileExistsError
            If the target database exists and overwrite=False.
        IOError
            If the restore cannot be completed due to permission or disk issues.
        
        Example
        -------
        >>> db.restore("backups/trades_backup_2024_01_01.db")
        >>> db.restore("backups/trades_backup.db", overwrite=True)
        """
        backup_path = Path(backup_path)
        
        # Ensure backup file exists
        if not backup_path.exists():
            raise FileNotFoundError(f"Backup file not found: {backup_path}")
        
        # Check if target exists and overwrite is False
        if self.db_path.exists() and not overwrite:
            raise FileExistsError(
                f"Database file already exists: {self.db_path}. "
                f"Use overwrite=True to replace it."
            )
        
        # Close connection before restoring
        was_open = self._conn is not None
        if was_open:
            self.close()
        
        try:
            # Copy the backup file to the database location
            shutil.copy2(backup_path, self.db_path)
            
            # Invalidate cache since database has changed
            self._invalidate_cache()
            
            log.info("Database restored: %s -> %s", backup_path, self.db_path)
        finally:
            # Reopen connection
            self._get_connection()
    
    def backup_to_s3(
        self,
        s3_uri: str,
        aws_access_key_id: Optional[str] = None,
        aws_secret_access_key: Optional[str] = None,
        region_name: Optional[str] = None,
        bucket_name: Optional[str] = None,
        key: Optional[str] = None,
    ) -> None:
        """
        Backup the database to Amazon S3.
        
        This method uploads the database file to Amazon S3. You can provide
        credentials directly or rely on environment variables/AWS credentials chain.
        
        Parameters
        ----------
        s3_uri : str
            S3 URI in format s3://bucket-name/key or just the full S3 path.
        aws_access_key_id : str, optional
            AWS access key ID. If not provided, uses default credential chain.
        aws_secret_access_key : str, optional
            AWS secret access key. If not provided, uses default credential chain.
        region_name : str, optional
            AWS region name (e.g., "us-east-1").
        bucket_name : str, optional
            S3 bucket name (alternative to s3_uri).
        key : str, optional
            S3 object key (alternative to s3_uri).
        
        Raises
        ------
        ImportError
            If boto3 is not installed.
        FileNotFoundError
            If the database file doesn't exist.
        Exception
            If the upload fails.
        
        Example
        -------
        >>> db.backup_to_s3("s3://my-bucket/backups/trades.db")
        >>> db.backup_to_s3("s3://my-bucket/backups/trades_2024_01_01.db",
        ...                  region_name="us-east-1")
        """
        # Ensure source database exists
        if not self.db_path.exists():
            raise FileNotFoundError(f"Source database not found: {self.db_path}")
        
        # Parse S3 URI
        if s3_uri.startswith("s3://"):
            # Parse s3://bucket/key format
            uri_parts = s3_uri[5:].split("/", 1)
            bucket = uri_parts[0]
            object_key = uri_parts[1] if len(uri_parts) > 1 else Path(self.db_path).name
        else:
            # Use provided bucket_name and key
            bucket = bucket_name
            object_key = key or Path(self.db_path).name
        
        if not bucket:
            raise ValueError("Bucket name must be provided via s3_uri or bucket_name parameter")
        
        try:
            import boto3
            from botocore.exceptions import ClientError
        except ImportError as e:
            raise ImportError(
                "boto3 is required for S3 backup. "
                "Install it with: pip install boto3"
            ) from e
        
        # Close connection before uploading to ensure data integrity
        was_open = self._conn is not None
        if was_open:
            self.close()
        
        try:
            # Create S3 client
            session_kwargs = {}
            if aws_access_key_id and aws_secret_access_key:
                session_kwargs["aws_access_key_id"] = aws_access_key_id
                session_kwargs["aws_secret_access_key"] = aws_secret_access_key
            if region_name:
                session_kwargs["region_name"] = region_name
            
            s3_client = boto3.client("s3", **session_kwargs)
            
            # Upload file
            s3_client.upload_file(str(self.db_path), bucket, object_key)
            
            log.info("Database backup uploaded to S3: s3://%s/%s", bucket, object_key)
        except ClientError as e:
            log.error("S3 backup failed: %s", e)
            raise
        finally:
            # Reopen connection if it was open
            if was_open:
                self._get_connection()
    
    def backup_to_gcs(
        self,
        gcs_uri: str,
        credentials_path: Optional[str] = None,
        project_id: Optional[str] = None,
        bucket_name: Optional[str] = None,
        blob_name: Optional[str] = None,
    ) -> None:
        """
        Backup the database to Google Cloud Storage.
        
        This method uploads the database file to Google Cloud Storage. You can provide
        credentials via a service account JSON file or rely on Application Default Credentials.
        
        Parameters
        ----------
        gcs_uri : str
            GCS URI in format gs://bucket-name/blob-name or just the full GCS path.
        credentials_path : str, optional
            Path to service account JSON credentials file.
        project_id : str, optional
            GCP project ID.
        bucket_name : str, optional
            GCS bucket name (alternative to gcs_uri).
        blob_name : str, optional
            GCS blob name (alternative to gcs_uri).
        
        Raises
        ------
        ImportError
            If google-cloud-storage is not installed.
        FileNotFoundError
            If the database file or credentials file doesn't exist.
        Exception
            If the upload fails.
        
        Example
        -------
        >>> db.backup_to_gcs("gs://my-bucket/backups/trades.db")
        >>> db.backup_to_gcs("gs://my-bucket/backups/trades_2024_01_01.db",
        ...                  credentials_path="service_account.json",
        ...                  project_id="my-project")
        """
        # Ensure source database exists
        if not self.db_path.exists():
            raise FileNotFoundError(f"Source database not found: {self.db_path}")
        
        # Parse GCS URI
        if gcs_uri.startswith("gs://"):
            # Parse gs://bucket/blob format
            uri_parts = gcs_uri[5:].split("/", 1)
            bucket = uri_parts[0]
            blob = uri_parts[1] if len(uri_parts) > 1 else Path(self.db_path).name
        else:
            # Use provided bucket_name and blob_name
            bucket = bucket_name
            blob = blob_name or Path(self.db_path).name
        
        if not bucket:
            raise ValueError("Bucket name must be provided via gcs_uri or bucket_name parameter")
        
        try:
            from google.cloud import storage
        except ImportError as e:
            raise ImportError(
                "google-cloud-storage is required for GCS backup. "
                "Install it with: pip install google-cloud-storage"
            ) from e
        
        # Close connection before uploading to ensure data integrity
        was_open = self._conn is not None
        if was_open:
            self.close()
        
        try:
            # Create GCS client
            client_kwargs = {}
            if credentials_path:
                credentials_path_obj = Path(credentials_path)
                if not credentials_path_obj.exists():
                    raise FileNotFoundError(f"Credentials file not found: {credentials_path}")
                from google.oauth2 import service_account
                credentials = service_account.Credentials.from_service_account_file(
                    str(credentials_path_obj)
                )
                client_kwargs["credentials"] = credentials
            if project_id:
                client_kwargs["project"] = project_id
            
            client = storage.Client(**client_kwargs)
            
            # Get bucket and upload
            bucket_obj = client.bucket(bucket)
            blob_obj = bucket_obj.blob(blob)
            blob_obj.upload_from_filename(str(self.db_path))
            
            log.info("Database backup uploaded to GCS: gs://%s/%s", bucket, blob)
        except Exception as e:
            log.error("GCS backup failed: %s", e)
            raise
        finally:
            # Reopen connection if it was open
            if was_open:
                self._get_connection()
    
    def delete_trade(self, trade_id: int) -> bool:
        """
        Delete a trade by ID.
        
        Parameters
        ----------
        trade_id : int
            ID of the trade to delete.
        
        Returns
        -------
        bool
            True if trade was deleted, False if not found.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM trades WHERE id = ?", (trade_id,))
        conn.commit()
        
        deleted = cursor.rowcount > 0
        if deleted:
            # Invalidate cache on write
            self._invalidate_cache()
            
            # Trigger trade_deleted hooks if streaming is enabled
            if self._streaming_enabled:
                self._trigger_trade_deleted_hooks(trade_id)
            
            log.debug("Trade deleted: ID=%d", trade_id)
        return deleted
    
    def clear_all_trades(self) -> None:
        """Delete all trades from the database."""
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM trades")
        conn.commit()
        
        # Invalidate cache on write
        self._invalidate_cache()
        
        log.info("All trades cleared from database")
    
    # Event Hooks for Real-time Synchronization
    
    def on_trade_saved(self, callback: Callable[[TradeRecord], None]) -> Callable[[TradeRecord], None]:
        """
        Register a callback to be called when a trade is saved.
        
        Parameters
        ----------
        callback : Callable[[TradeRecord], None]
            Function that takes a TradeRecord as argument.
        
        Returns
        -------
        Callable[[TradeRecord], None]
            The same callback function (allows decorator usage).
        
        Example
        -------
        >>> @db.on_trade_saved
        >>> def handle_trade_saved(trade: TradeRecord):
        ...     print(f"Trade saved: {trade.market_slug}")
        """
        with self._hooks_lock:
            self._trade_saved_hooks.append(callback)
        log.debug("Registered trade_saved callback: %s", callback.__name__)
        return callback
    
    def on_trade_updated(self, callback: Callable[[int, Dict[str, Any]], None]) -> Callable[[int, Dict[str, Any]], None]:
        """
        Register a callback to be called when a trade is updated.
        
        Parameters
        ----------
        callback : Callable[[int, Dict[str, Any]], None]
            Function that takes trade_id and changes dict as arguments.
        
        Returns
        -------
        Callable[[int, Dict[str, Any]], None]
            The same callback function (allows decorator usage).
        
        Example
        -------
        >>> @db.on_trade_updated
        >>> def handle_trade_updated(trade_id: int, changes: Dict[str, Any]):
        ...     print(f"Trade {trade_id} updated: {changes}")
        """
        with self._hooks_lock:
            self._trade_updated_hooks.append(callback)
        log.debug("Registered trade_updated callback: %s", callback.__name__)
        return callback
    
    def on_trade_deleted(self, callback: Callable[[int], None]) -> Callable[[int], None]:
        """
        Register a callback to be called when a trade is deleted.
        
        Parameters
        ----------
        callback : Callable[[int], None]
            Function that takes trade_id as argument.
        
        Returns
        -------
        Callable[[int], None]
            The same callback function (allows decorator usage).
        
        Example
        -------
        >>> @db.on_trade_deleted
        >>> def handle_trade_deleted(trade_id: int):
        ...     print(f"Trade {trade_id} deleted")
        """
        with self._hooks_lock:
            self._trade_deleted_hooks.append(callback)
        log.debug("Registered trade_deleted callback: %s", callback.__name__)
        return callback
    
    def remove_trade_saved_hook(self, callback: Callable[[TradeRecord], None]) -> None:
        """
        Remove a registered trade_saved callback.
        
        Parameters
        ----------
        callback : Callable[[TradeRecord], None]
            The callback function to remove.
        """
        with self._hooks_lock:
            if callback in self._trade_saved_hooks:
                self._trade_saved_hooks.remove(callback)
                log.debug("Removed trade_saved callback: %s", callback.__name__)
    
    def remove_trade_updated_hook(self, callback: Callable[[int, Dict[str, Any]], None]) -> None:
        """
        Remove a registered trade_updated callback.
        
        Parameters
        ----------
        callback : Callable[[int, Dict[str, Any]], None]
            The callback function to remove.
        """
        with self._hooks_lock:
            if callback in self._trade_updated_hooks:
                self._trade_updated_hooks.remove(callback)
                log.debug("Removed trade_updated callback: %s", callback.__name__)
    
    def remove_trade_deleted_hook(self, callback: Callable[[int], None]) -> None:
        """
        Remove a registered trade_deleted callback.
        
        Parameters
        ----------
        callback : Callable[[int], None]
            The callback function to remove.
        """
        with self._hooks_lock:
            if callback in self._trade_deleted_hooks:
                self._trade_deleted_hooks.remove(callback)
                log.debug("Removed trade_deleted callback: %s", callback.__name__)
    
    def _trigger_trade_saved_hooks(self, trade: TradeRecord) -> None:
        """
        Trigger all registered trade_saved callbacks.
        
        Parameters
        ----------
        trade : TradeRecord
            The trade that was saved.
        """
        with self._hooks_lock:
            for callback in self._trade_saved_hooks:
                try:
                    callback(trade)
                except Exception as e:
                    log.error("Error in trade_saved callback %s: %s", callback.__name__, e)
    
    def _trigger_trade_updated_hooks(self, trade_id: int, changes: Dict[str, Any]) -> None:
        """
        Trigger all registered trade_updated callbacks.
        
        Parameters
        ----------
        trade_id : int
            The ID of the trade that was updated.
        changes : Dict[str, Any]
            Dictionary of changed fields.
        """
        with self._hooks_lock:
            for callback in self._trade_updated_hooks:
                try:
                    callback(trade_id, changes)
                except Exception as e:
                    log.error("Error in trade_updated callback %s: %s", callback.__name__, e)
    
    def _trigger_trade_deleted_hooks(self, trade_id: int) -> None:
        """
        Trigger all registered trade_deleted callbacks.
        
        Parameters
        ----------
        trade_id : int
            The ID of the trade that was deleted.
        """
        with self._hooks_lock:
            for callback in self._trade_deleted_hooks:
                try:
                    callback(trade_id)
                except Exception as e:
                    log.error("Error in trade_deleted callback %s: %s", callback.__name__, e)
    
    # Streaming Control
    
    def enable_streaming(self) -> None:
        """
        Enable real-time streaming of database events.
        
        When enabled, registered event hooks will be triggered
        on trade save, update, and delete operations.
        """
        self._streaming_enabled = True
        log.info("Real-time streaming enabled")
    
    def disable_streaming(self) -> None:
        """
        Disable real-time streaming of database events.
        
        When disabled, event hooks will not be triggered.
        """
        self._streaming_enabled = False
        log.info("Real-time streaming disabled")
    
    def is_streaming_enabled(self) -> bool:
        """
        Check if streaming is currently enabled.
        
        Returns
        -------
        bool
            True if streaming is enabled, False otherwise.
        """
        return self._streaming_enabled
    
    # Change Data Capture
    
    def get_recent_changes(self, limit: int = 100) -> List[Dict[str, Any]]:
        """
        Get recent database changes from the change log.
        
        This method queries the trades table ordered by created_at
        to simulate change tracking. For production use, consider
        implementing a dedicated change log table.
        
        Parameters
        ----------
        limit : int, optional
            Maximum number of changes to return (default: 100).
        
        Returns
        -------
        List[Dict[str, Any]]
            List of recent changes with metadata.
        """
        conn = self._get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, market_slug, market_id, side, outcome, pnl, 
                   timestamp, created_at
            FROM trades
            ORDER BY created_at DESC
            LIMIT ?
        """, (limit,))
        
        changes = []
        for row in cursor.fetchall():
            changes.append({
                "id": row["id"],
                "market_slug": row["market_slug"],
                "market_id": row["market_id"],
                "side": row["side"],
                "outcome": row["outcome"],
                "pnl": row["pnl"],
                "timestamp": row["timestamp"],
                "created_at": row["created_at"],
                "operation": "INSERT"  # Simulated operation type
            })
        
        log.debug("Retrieved %d recent changes", len(changes))
        return changes
    
    # Monitoring and Observability
    
    def get_metrics(self) -> DatabaseMetrics:
        """
        Get database performance and health metrics.
        
        Returns
        -------
        DatabaseMetrics
            Current database metrics including size, cache performance,
            query statistics, and configuration.
        
        Example
        -------
        >>> metrics = db.get_metrics()
        >>> print(f"Database size: {metrics.database_size_mb} MB")
        >>> print(f"Cache hit rate: {metrics.cache_hit_rate}%")
        """
        # Get total trades count
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM trades")
        total_trades = cursor.fetchone()[0]
        
        # Get database size
        database_size_bytes = self.db_path.stat().st_size if self.db_path.exists() else 0
        
        # Calculate cache hit rate
        total_cache_requests = self._cache_hits + self._cache_misses
        cache_hit_rate = (self._cache_hits / total_cache_requests) if total_cache_requests > 0 else 0.0
        
        # Calculate average query time
        avg_query_time_ms = (sum(self._query_times) / len(self._query_times)) if self._query_times else 0.0
        
        # Get WAL status
        cursor.execute("PRAGMA journal_mode")
        wal_mode = cursor.fetchone()[0]
        wal_enabled = wal_mode.upper() == "WAL"
        
        return DatabaseMetrics(
            total_trades=total_trades,
            database_size_bytes=database_size_bytes,
            cache_hit_rate=cache_hit_rate,
            cache_size=len(self._query_cache),
            query_count=self._query_count,
            slow_query_count=self._slow_query_count,
            avg_query_time_ms=avg_query_time_ms,
            connection_pool_size=1 if self._conn else 0,
            wal_enabled=wal_enabled,
            last_optimization=self._last_optimization,
        )
    
    def get_logs(
        self,
        level: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        operation: Optional[str] = None,
        limit: int = 100,
    ) -> List[LogEntry]:
        """
        Get structured logs with correlation IDs.
        
        Parameters
        ----------
        level : str, optional
            Filter by log level (e.g., "DEBUG", "INFO", "WARNING", "ERROR").
        start_date : datetime, optional
            Filter logs after this timestamp.
        end_date : datetime, optional
            Filter logs before this timestamp.
        operation : str, optional
            Filter by operation type (e.g., "save_trade", "load_trades").
        limit : int, optional
            Maximum number of log entries to return (default: 100).
        
        Returns
        -------
        List[LogEntry]
            Filtered log entries.
        
        Example
        -------
        >>> logs = db.get_logs(level="ERROR", limit=50)
        >>> for entry in logs:
        ...     print(f"{entry.timestamp}: {entry.message}")
        """
        with self._log_lock:
            filtered_logs = self._log_entries.copy()
        
        # Apply filters
        if level:
            filtered_logs = [log for log in filtered_logs if log.level == level.upper()]
        
        if start_date:
            filtered_logs = [log for log in filtered_logs if log.timestamp >= start_date]
        
        if end_date:
            filtered_logs = [log for log in filtered_logs if log.timestamp <= end_date]
        
        if operation:
            filtered_logs = [log for log in filtered_logs if log.operation == operation]
        
        # Sort by timestamp descending and limit
        filtered_logs.sort(key=lambda x: x.timestamp, reverse=True)
        return filtered_logs[:limit]
    
    def set_alert(
        self,
        name: str,
        metric: str,
        threshold: float,
        comparison: str = "gt",
        callback: Optional[Callable[[str, float, float], None]] = None,
    ) -> None:
        """
        Set up an alert rule for monitoring database metrics.
        
        Parameters
        ----------
        name : str
            Unique name for the alert rule.
        metric : str
            Metric to monitor (e.g., "slow_query_count", "avg_query_time_ms", "cache_hit_rate").
        threshold : float
            Threshold value for triggering the alert.
        comparison : str, optional
            Comparison operator: "gt" (greater than), "lt" (less than),
            "eq" (equal), "gte" (greater or equal), "lte" (less or equal) (default: "gt").
        callback : callable, optional
            Function to call when alert is triggered. Receives (name, metric_value, threshold).
        
        Raises
        ------
        ValueError
            If comparison is invalid.
        
        Example
        -------
        >>> def on_slow_query(name, value, threshold):
        ...     print(f"Alert {name}: Query time {value}ms exceeds threshold {threshold}ms")
        >>> db.set_alert("slow_query", "avg_query_time_ms", 1000.0, callback=on_slow_query)
        """
        valid_comparisons = {"gt", "lt", "eq", "gte", "lte"}
        if comparison not in valid_comparisons:
            raise ValueError(
                f"Invalid comparison '{comparison}'. Valid options: {valid_comparisons}"
            )
        
        with self._alert_lock:
            self._alert_rules[name] = AlertRule(
                name=name,
                metric=metric,
                threshold=threshold,
                comparison=comparison,
                enabled=True,
                callback=callback,
                last_triggered=None,
                trigger_count=0,
            )
        
        log.info("Alert rule set: %s on %s %s %s", name, metric, comparison, threshold)
    
    def remove_alert(self, name: str) -> None:
        """
        Remove an alert rule.
        
        Parameters
        ----------
        name : str
            Name of the alert rule to remove.
        """
        with self._alert_lock:
            if name in self._alert_rules:
                del self._alert_rules[name]
                log.info("Alert rule removed: %s", name)
    
    def get_alerts(self) -> Dict[str, Dict[str, Any]]:
        """
        Get all configured alert rules.
        
        Returns
        -------
        dict
            Dictionary of alert rules with their status.
        """
        with self._alert_lock:
            return {
                name: {
                    "metric": rule.metric,
                    "threshold": rule.threshold,
                    "comparison": rule.comparison,
                    "enabled": rule.enabled,
                    "last_triggered": rule.last_triggered.isoformat() if rule.last_triggered else None,
                    "trigger_count": rule.trigger_count,
                }
                for name, rule in self._alert_rules.items()
            }
    
    def check_alerts(self) -> None:
        """
        Check all alert rules and trigger callbacks if thresholds are exceeded.
        
        This method should be called periodically to evaluate alert conditions.
        """
        metrics = self.get_metrics()
        metric_values = metrics.to_dict()
        
        with self._alert_lock:
            for name, rule in self._alert_rules.items():
                if not rule.enabled:
                    continue
                
                # Get current metric value
                metric_value = metric_values.get(rule.metric)
                if metric_value is None:
                    continue
                
                # Check if threshold is exceeded
                triggered = False
                if rule.comparison == "gt" and metric_value > rule.threshold:
                    triggered = True
                elif rule.comparison == "lt" and metric_value < rule.threshold:
                    triggered = True
                elif rule.comparison == "eq" and metric_value == rule.threshold:
                    triggered = True
                elif rule.comparison == "gte" and metric_value >= rule.threshold:
                    triggered = True
                elif rule.comparison == "lte" and metric_value <= rule.threshold:
                    triggered = True
                
                if triggered:
                    rule.last_triggered = datetime.now(timezone.utc)
                    rule.trigger_count += 1
                    
                    # Call callback if provided
                    if rule.callback:
                        try:
                            rule.callback(name, metric_value, rule.threshold)
                        except Exception as e:
                            log.error("Error in alert callback for %s: %s", name, e)
                    
                    log.warning(
                        "Alert triggered: %s - %s = %s (threshold: %s)",
                        name, rule.metric, metric_value, rule.threshold
                    )
    
    @contextmanager
    def _track_query(self, operation: str):
        """
        Context manager for tracking query performance.
        
        Parameters
        ----------
        operation : str
            Name of the operation being tracked.
        """
        correlation_id = self._get_correlation_id()
        start_time = time.perf_counter()
        
        try:
            yield
        finally:
            duration_ms = (time.perf_counter() - start_time) * 1000
            self._query_count += 1
            
            # Track query time
            self._query_times.append(duration_ms)
            if len(self._query_times) > self._max_query_times:
                self._query_times.pop(0)
            
            # Track slow queries
            if duration_ms > self._slow_query_threshold_ms:
                self._slow_query_count += 1
                self._add_log_entry(
                    level="WARNING",
                    message=f"Slow query detected: {operation}",
                    operation=operation,
                    duration_ms=duration_ms,
                )
            
            # Log query completion
            self._add_log_entry(
                level="DEBUG",
                message=f"Query completed: {operation}",
                operation=operation,
                duration_ms=duration_ms,
            )
    
    def _add_log_entry(
        self,
        level: str,
        message: str,
        operation: Optional[str] = None,
        duration_ms: Optional[float] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> None:
        """
        Add a structured log entry.
        
        Parameters
        ----------
        level : str
            Log level (DEBUG, INFO, WARNING, ERROR).
        message : str
            Log message.
        operation : str, optional
            Operation being performed.
        duration_ms : float, optional
            Operation duration in milliseconds.
        metadata : dict, optional
            Additional metadata.
        """
        correlation_id = self._get_correlation_id()
        entry = LogEntry(
            correlation_id=correlation_id,
            timestamp=datetime.now(timezone.utc),
            level=level.upper(),
            message=message,
            operation=operation,
            duration_ms=duration_ms,
            metadata=metadata or {},
        )
        
        with self._log_lock:
            self._log_entries.append(entry)
            if len(self._log_entries) > self._max_log_entries:
                self._log_entries.pop(0)
    
    def set_correlation_id(self, correlation_id: str) -> None:
        """
        Set a correlation ID for the current operation context.
        
        Parameters
        ----------
        correlation_id : str
            Correlation ID to use for logging.
        
        Example
        -------
        >>> db.set_correlation_id("req-12345")
        >>> db.save_trade(...)  # All logs will use this correlation ID
        """
        with self._correlation_lock:
            self._current_correlation_id = correlation_id
    
    def clear_correlation_id(self) -> None:
        """Clear the current correlation ID."""
        with self._correlation_lock:
            self._current_correlation_id = None
    
    def _get_correlation_id(self) -> str:
        """
        Get the current correlation ID, generating one if not set.
        
        Returns
        -------
        str
            Current correlation ID.
        """
        with self._correlation_lock:
            if self._current_correlation_id is None:
                self._current_correlation_id = str(uuid.uuid4())
            return self._current_correlation_id
    
    @contextmanager
    def operation_context(self, operation_name: str):
        """
        Context manager for tracking operations with correlation IDs.
        
        Parameters
        ----------
        operation_name : str
            Name of the operation.
        
        Example
        -------
        >>> with db.operation_context("batch_import"):
        ...     db.save_trades_bulk(trades)
        """
        old_correlation_id = self._current_correlation_id
        self.set_correlation_id(str(uuid.uuid4()))
        
        self._add_log_entry(
            level="INFO",
            message=f"Operation started: {operation_name}",
            operation=operation_name,
        )
        
        start_time = time.perf_counter()
        
        try:
            yield
            duration_ms = (time.perf_counter() - start_time) * 1000
            self._add_log_entry(
                level="INFO",
                message=f"Operation completed: {operation_name}",
                operation=operation_name,
                duration_ms=duration_ms,
            )
        except Exception as e:
            duration_ms = (time.perf_counter() - start_time) * 1000
            self._add_log_entry(
                level="ERROR",
                message=f"Operation failed: {operation_name} - {str(e)}",
                operation=operation_name,
                duration_ms=duration_ms,
                metadata={"error": str(e), "error_type": type(e).__name__},
            )
            raise
        finally:
            self._current_correlation_id = old_correlation_id
    
    # Security Features
    
    def enable_encryption(self, key: Optional[bytes] = None, password: Optional[str] = None, fields: Optional[List[str]] = None) -> None:
        """
        Enable at-rest encryption for the database.
        
        Parameters
        ----------
        key : bytes, optional
            32-byte URL-safe base64-encoded encryption key.
        password : str, optional
            Password to derive encryption key from.
        fields : list of str, optional
            List of field names to encrypt (default: ["market_id"]).
        
        Raises
        ------
        ImportError
            If cryptography library is not installed.
        
        Example
        -------
        >>> db.enable_encryption(password="my_secure_password")
        >>> db.enable_encryption(key=DatabaseEncryption.generate_key(), fields=["market_id"])
        """
        self._encryption = DatabaseEncryption(key=key, password=password)
        self._encryption_fields = fields or ["market_id"]
        log.info("Encryption enabled for fields: %s", self._encryption_fields)
    
    def disable_encryption(self) -> None:
        """Disable encryption (data will be stored in plaintext)."""
        if self._encryption:
            self._encryption.disable()
        log.info("Encryption disabled")
    
    def is_encryption_enabled(self) -> bool:
        """Check if encryption is enabled."""
        return self._encryption is not None and self._encryption.is_enabled()
    
    def set_auth_method(self, method: str) -> None:
        """
        Set authentication method.
        
        Parameters
        ----------
        method : str
            Authentication method: "none", "api_key", or "jwt".
        
        Example
        -------
        >>> db.set_auth_method("api_key")
        >>> db.set_auth_method("jwt")
        """
        auth_method = AuthMethod(method.lower())
        self._auth_manager.set_method(auth_method)
        log.info("Authentication method set to: %s", method)
    
    def get_auth_method(self) -> str:
        """Get current authentication method."""
        return self._auth_manager.get_method().value
    
    def add_user(
        self,
        user_id: str,
        username: str,
        roles: List[str],
        api_key: Optional[str] = None,
        jwt_secret: Optional[str] = None,
    ) -> None:
        """
        Add a user to the authentication system.
        
        Parameters
        ----------
        user_id : str
            Unique user identifier.
        username : str
            Username.
        roles : list of str
            List of role names for the user.
        api_key : str, optional
            API key for the user (auto-generated if not provided).
        jwt_secret : str, optional
            JWT secret for the user.
        
        Example
        -------
        >>> db.add_user("user1", "trader", ["trader"])
        >>> db.add_user("user2", "analyst", ["analyst"], api_key="pk_...")
        """
        if api_key is None and self._auth_manager.get_method() == AuthMethod.API_KEY:
            api_key = self._auth_manager.generate_api_key()
        
        self._auth_manager.add_user(user_id, username, roles, api_key, jwt_secret)
    
    def remove_user(self, user_id: str) -> None:
        """
        Remove a user from the authentication system.
        
        Parameters
        ----------
        user_id : str
            User identifier to remove.
        """
        self._auth_manager.remove_user(user_id)
    
    def authenticate(self, credential: str, user_id: Optional[str] = None) -> bool:
        """
        Authenticate a user using the configured authentication method.
        
        Parameters
        ----------
        credential : str
            Authentication credential (API key or JWT token).
        user_id : str, optional
            User ID (required for JWT authentication).
        
        Returns
        -------
        bool
            True if authentication successful, False otherwise.
        
        Example
        -------
        >>> if db.authenticate("pk_abc123"):
        ...     print("Authenticated!")
        >>> if db.authenticate("jwt_token", user_id="user1"):
        ...     print("Authenticated!")
        """
        method = self._auth_manager.get_method()
        
        if method == AuthMethod.API_KEY:
            authenticated_user_id = self._auth_manager.validate_api_key(credential)
            if authenticated_user_id:
                self._current_user_id = authenticated_user_id
                user = self._auth_manager.get_user(authenticated_user_id)
                if user:
                    self._current_roles = user.roles
                return True
            return False
        
        elif method == AuthMethod.JWT:
            if user_id is None:
                raise ValueError("user_id is required for JWT authentication")
            if self._auth_manager.validate_jwt_token(credential, user_id):
                self._current_user_id = user_id
                user = self._auth_manager.get_user(user_id)
                if user:
                    self._current_roles = user.roles
                return True
            return False
        
        elif method == AuthMethod.NONE:
            return True
        
        return False
    
    def get_current_user(self) -> Optional[str]:
        """Get the currently authenticated user ID."""
        return self._current_user_id
    
    def get_current_roles(self) -> Set[str]:
        """Get the roles of the currently authenticated user."""
        return self._current_roles.copy()
    
    def add_role(self, name: str, permissions: List[str], description: Optional[str] = None) -> None:
        """
        Add a custom role.
        
        Parameters
        ----------
        name : str
            Role name.
        permissions : list of str
            List of permissions for the role.
        description : str, optional
            Role description.
        
        Example
        -------
        >>> db.add_role("manager", ["read", "write", "export"], "Can read, write, and export")
        """
        role = Role(name=name, permissions=set(permissions), description=description)
        self._authz_manager.add_role(role)
    
    def remove_role(self, role_name: str) -> None:
        """
        Remove a role.
        
        Parameters
        ----------
        role_name : str
            Role name to remove.
        """
        self._authz_manager.remove_role(role_name)
    
    def check_permission(self, permission: str) -> bool:
        """
        Check if the current user has a specific permission.
        
        Parameters
        ----------
        permission : str
            Permission to check (e.g., "read", "write", "delete").
        
        Returns
        -------
        bool
            True if user has permission, False otherwise.
        
        Example
        -------
        >>> if db.check_permission("write"):
        ...     db.save_trade(...)
        """
        if not self._auth_manager.is_enabled():
            return True
        return self._authz_manager.check_permission(self._current_roles, permission)
    
    def require_permission(self, permission: str) -> None:
        """
        Require a specific permission for the current operation.
        
        Raises PermissionError if the current user doesn't have the permission.
        
        Parameters
        ----------
        permission : str
            Required permission.
        
        Raises
        ------
        PermissionError
            If user doesn't have the required permission.
        
        Example
        -------
        >>> db.require_permission("delete")
        >>> db.delete_trade(trade_id)
        """
        if not self.check_permission(permission):
            raise PermissionError(
                f"User '{self._current_user_id}' with roles {self._current_roles} "
                f"does not have permission '{permission}'"
            )
    
    def add_masking_rule(self, field_name: str, mask_char: str = "*", show_first: int = 0, show_last: int = 0, mask_all: bool = False) -> None:
        """
        Add a data masking rule for a field.
        
        Parameters
        ----------
        field_name : str
            Field name to mask.
        mask_char : str, optional
            Character to use for masking (default: "*").
        show_first : int, optional
            Number of characters to show at the beginning (default: 0).
        show_last : int, optional
            Number of characters to show at the end (default: 0).
        mask_all : bool, optional
            Mask the entire field (default: False).
        
        Example
        -------
        >>> db.add_masking_rule("market_id", show_first=4, show_last=4)
        >>> db.add_masking_rule("secret", mask_all=True)
        """
        rule = MaskingRule(
            field_name=field_name,
            mask_char=mask_char,
            show_first=show_first,
            show_last=show_last,
            mask_all=mask_all,
        )
        self._data_masker.add_rule(rule)
    
    def remove_masking_rule(self, field_name: str) -> None:
        """
        Remove a masking rule.
        
        Parameters
        ----------
        field_name : str
            Field name to remove rule for.
        """
        self._data_masker.remove_rule(field_name)
    
    def enable_masking(self) -> None:
        """Enable data masking."""
        self._data_masker.enable()
    
    def disable_masking(self) -> None:
        """Disable data masking."""
        self._data_masker.disable()
    
    def is_masking_enabled(self) -> bool:
        """Check if data masking is enabled."""
        return self._data_masker.is_enabled()
    
    def mask_trade_record(self, trade: TradeRecord) -> Dict[str, Any]:
        """
        Apply masking to a trade record.
        
        Parameters
        ----------
        trade : TradeRecord
            Trade record to mask.
        
        Returns
        -------
        dict
            Masked trade record as dictionary.
        """
        record_dict = trade.to_dict()
        return self._data_masker.mask_record(record_dict)
    
    def close(self) -> None:
        """Close the database connection and cleanup resources."""
        # Stop optimization thread if running
        if self._optimization_thread is not None:
            self._optimization_stop_event.set()
            self._optimization_thread.join(timeout=5)
            self._optimization_thread = None
        
        # Close all connections in pool
        while not self._connection_pool.empty():
            try:
                conn = self._connection_pool.get_nowait()
                conn.close()
            except Empty:
                break
        
        # Close legacy connection if exists
        if self._conn is not None:
            self._conn.close()
            self._conn = None
        
        log.debug("Database connection closed")
    
    def __enter__(self):
        """Context manager entry."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close()
        return False
    
    def _row_to_trade_record(self, row: sqlite3.Row) -> TradeRecord:
        """Convert a database row to a TradeRecord."""
        timestamp = datetime.fromisoformat(row["timestamp"])
        if timestamp.tzinfo is None:
            timestamp = timestamp.replace(tzinfo=timezone.utc)
        
        # Handle market_session - may not exist in older databases
        # sqlite3.Row doesn't have .get(), so we need to check key existence
        try:
            market_session = row["market_session"]
        except (KeyError, IndexError):
            market_session = None
        
        return TradeRecord(
            id=row["id"],
            market_slug=row["market_slug"],
            market_id=row["market_id"],
            side=row["side"],
            entry_price=row["entry_price"],
            exit_price=row["exit_price"],
            amount=row["amount"],
            shares=row["shares"],
            fee=row["fee"],
            outcome=row["outcome"],
            pnl=row["pnl"],
            timestamp=timestamp,
            market_session=market_session,
        )
