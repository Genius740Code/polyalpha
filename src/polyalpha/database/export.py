from __future__ import annotations

import csv
import json
import logging
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, List, Dict, Any

from ..utils.logging_utils import mask_address

log = logging.getLogger(__name__)


class DatabaseExporter:
    def __init__(self, db_instance):
        self._db = db_instance

    def export_csv(self, filepath: str | Path, filters: Optional[Dict[str, Any]] = None) -> None:
        filepath = Path(filepath)
        trades = self._db.load_trades(filters=filters)
        if not trades:
            log.warning("No trades to export to CSV")
            return
        fieldnames = ["id", "market_slug", "market_id", "side", "entry_price",
                       "exit_price", "amount", "shares", "fee", "outcome", "pnl", "timestamp"]
        with filepath.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for trade in trades:
                row = {"id": trade.id, "market_slug": trade.market_slug,
                       "market_id": trade.market_id, "side": trade.side,
                       "entry_price": trade.entry_price, "exit_price": trade.exit_price,
                       "amount": trade.amount, "shares": trade.shares, "fee": trade.fee,
                       "outcome": trade.outcome, "pnl": trade.pnl,
                       "timestamp": trade.timestamp.isoformat()}
                writer.writerow(row)
        log.info("Exported %d trades to CSV: %s", len(trades), filepath)

    def export_json(self, filepath: str | Path, filters: Optional[Dict[str, Any]] = None) -> None:
        filepath = Path(filepath)
        trades = self._db.load_trades(filters=filters)
        if not trades:
            log.warning("No trades to export to JSON")
            return
        export_data = {
            "metadata": {
                "export_timestamp": datetime.now(timezone.utc).isoformat(),
                "total_trades": len(trades),
                "database_path": str(self._db.db_path),
            },
            "trades": [trade.to_dict() for trade in trades]
        }
        with filepath.open("w", encoding="utf-8") as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        log.info("Exported %d trades to JSON: %s", len(trades), filepath)

    def export_parquet(self, filepath: str | Path, filters: Optional[Dict[str, Any]] = None) -> None:
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError as e:
            raise ImportError("pyarrow is required for Parquet export. Install it with: pip install pyarrow") from e
        filepath = Path(filepath)
        trades = self._db.load_trades(filters=filters)
        if not trades:
            log.warning("No trades to export to Parquet")
            return
        data = [trade.to_dict() for trade in trades]
        table = pa.Table.from_pylist(data)
        pq.write_table(table, filepath)
        log.info("Exported %d trades to Parquet: %s", len(trades), filepath)

    def export_excel(self, filepath: str | Path, filters: Optional[Dict[str, Any]] = None) -> None:
        try:
            from openpyxl import Workbook
        except ImportError as e:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl") from e
        filepath = Path(filepath)
        trades = self._db.load_trades(filters=filters)
        if not trades:
            log.warning("No trades to export to Excel")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        headers = ["ID", "Market Slug", "Market ID", "Side", "Entry Price",
                    "Exit Price", "Amount", "Shares", "Fee", "Outcome", "P&L", "Timestamp"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        for row_num, trade in enumerate(trades, 2):
            ws.cell(row=row_num, column=1, value=trade.id)
            ws.cell(row=row_num, column=2, value=trade.market_slug)
            ws.cell(row=row_num, column=3, value=trade.market_id)
            ws.cell(row=row_num, column=4, value=trade.side)
            ws.cell(row=row_num, column=5, value=trade.entry_price)
            ws.cell(row=row_num, column=6, value=trade.exit_price)
            ws.cell(row=row_num, column=7, value=trade.amount)
            ws.cell(row=row_num, column=8, value=trade.shares)
            ws.cell(row=row_num, column=9, value=trade.fee)
            ws.cell(row=row_num, column=10, value=trade.outcome)
            ws.cell(row=row_num, column=11, value=trade.pnl)
            ws.cell(row=row_num, column=12, value=trade.timestamp.isoformat())
        wb.save(filepath)
        log.info("Exported %d trades to Excel: %s", len(trades), filepath)


class DatabaseBackup:
    def __init__(self, db_instance):
        self._db = db_instance

    def backup(self, backup_path: str | Path) -> None:
        backup_path = Path(backup_path)
        if not self._db.db_path.exists():
            raise FileNotFoundError(f"Source database not found: {self._db.db_path}")
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        self._db.close()
        try:
            shutil.copy2(self._db.db_path, backup_path)
            log.info("Database backup created: %s -> %s", self._db.db_path, backup_path)
        finally:
            self._db._conn_mgr._initialize_db()

    def restore(self, backup_path: str | Path, overwrite: bool = False) -> None:
        backup_path = Path(backup_path)
        if not backup_path.exists():
            raise FileNotFoundError(f"Backup file not found: {backup_path}")
        if self._db.db_path.exists() and not overwrite:
            raise FileExistsError(
                f"Database file already exists: {self._db.db_path}. Use overwrite=True to replace it."
            )
        self._db.close()
        try:
            shutil.copy2(backup_path, self._db.db_path)
            self._db._invalidate_cache()
            log.info("Database restored: %s -> %s", backup_path, self._db.db_path)
        finally:
            self._db._conn_mgr._initialize_db()

    def backup_to_s3(
        self,
        s3_uri: str,
        aws_access_key_id: Optional[str] = None,
        aws_secret_access_key: Optional[str] = None,
        region_name: Optional[str] = None,
        bucket_name: Optional[str] = None,
        key: Optional[str] = None,
    ) -> None:
        if not self._db.db_path.exists():
            raise FileNotFoundError(f"Source database not found: {self._db.db_path}")
        if s3_uri.startswith("s3://"):
            uri_parts = s3_uri[5:].split("/", 1)
            bucket = uri_parts[0]
            object_key = uri_parts[1] if len(uri_parts) > 1 else Path(self._db.db_path).name
        else:
            bucket = bucket_name
            object_key = key or Path(self._db.db_path).name
        if not bucket:
            raise ValueError("Bucket name must be provided via s3_uri or bucket_name parameter")
        try:
            import boto3
            from botocore.exceptions import ClientError
        except ImportError as e:
            raise ImportError("boto3 is required for S3 backup. Install it with: pip install boto3") from e
        self._db.close()
        try:
            session_kwargs = {}
            if aws_access_key_id and aws_secret_access_key:
                session_kwargs["aws_access_key_id"] = aws_access_key_id
                session_kwargs["aws_secret_access_key"] = aws_secret_access_key
            if region_name:
                session_kwargs["region_name"] = region_name
            s3_client = boto3.client("s3", **session_kwargs)
            s3_client.upload_file(str(self._db.db_path), bucket, object_key)
            log.info("Database backup uploaded to S3: s3://%s/%s", bucket, mask_address(object_key))
        except Exception as e:
            log.error("S3 backup failed: %s", e)
            raise
        finally:
            self._db._get_connection()

    def backup_to_gcs(
        self,
        gcs_uri: str,
        credentials_path: Optional[str] = None,
        project_id: Optional[str] = None,
        bucket_name: Optional[str] = None,
        blob_name: Optional[str] = None,
    ) -> None:
        if not self._db.db_path.exists():
            raise FileNotFoundError(f"Source database not found: {self._db.db_path}")
        if gcs_uri.startswith("gs://"):
            uri_parts = gcs_uri[5:].split("/", 1)
            bucket = uri_parts[0]
            blob = uri_parts[1] if len(uri_parts) > 1 else Path(self._db.db_path).name
        else:
            bucket = bucket_name
            blob = blob_name or Path(self._db.db_path).name
        if not bucket:
            raise ValueError("Bucket name must be provided via gcs_uri or bucket_name parameter")
        try:
            from google.cloud import storage
        except ImportError as e:
            raise ImportError("google-cloud-storage is required for GCS backup. Install it with: pip install google-cloud-storage") from e
        self._db.close()
        try:
            client_kwargs = {}
            if credentials_path:
                credentials_path_obj = Path(credentials_path)
                if not credentials_path_obj.exists():
                    raise FileNotFoundError(f"Credentials file not found: {credentials_path}")
                from google.oauth2 import service_account
                credentials = service_account.Credentials.from_service_account_file(str(credentials_path_obj))
                client_kwargs["credentials"] = credentials
            if project_id:
                client_kwargs["project"] = project_id
            client = storage.Client(**client_kwargs)
            bucket_obj = client.bucket(bucket)
            blob_obj = bucket_obj.blob(blob)
            blob_obj.upload_from_filename(str(self._db.db_path))
            log.info("Database backup uploaded to GCS: gs://%s/%s", bucket, blob)
        except Exception as e:
            log.error("GCS backup failed: %s", e)
            raise
        finally:
            self._db._get_connection()
